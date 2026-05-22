#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
强制执行组 - 裁定信息提取与导出模块

功能：
1. 批量处理裁定PDF，提取信息
2. 与台账（非诉表格）进行责令号匹配
3. 导出合并后Excel（原表字段 + OCR识别字段）
   - Sheet1: 匹配数据（台账中有且PDF中有）
   - Sheet2: PDF独有数据（台账中没有但PDF中有）
"""

import json
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Any, Set
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from core.paths import ROOT

from core.config_loader import load_config
from enforcement.extractor import RulingPDFExtractor, RulingInfo, extract_ruling_from_pdf, chinese_date_to_arabic
from enforcement.product import (
    EnforcementCaseRegistry,
    load_enforcement_cases,
    normalize_notice_for_match,
)

_cfg = load_config()
_enforcement_cfg = _cfg.raw_config.get('enforcement', {})
_paths_cfg = _enforcement_cfg.get('paths', {})


def build_output_excel(
    registry: EnforcementCaseRegistry,
    pdf_results: Dict[str, RulingInfo],
    output_path: Path,
) -> Path:
    """
    构建输出Excel：合并非诉表格字段与OCR识别字段

    输出两个Sheet：
      - Sheet1 "匹配数据": 台账中有且匹配到PDF的
      - Sheet2 "PDF独有": 台账中没有但PDF中有的责令号

    输出列：
      区号 | 行政审查案号 | 责令号 | 被执行人 | 职工姓名 | 金额 | 法官/法官助理 | 执行时间 | 裁定结果 | 备注
    """
    # 调试信息
    print(f"[DEBUG] build_output_excel: 台账行数={len(registry.cases)}, PDF数={len(pdf_results)}")
    for case in registry.cases:
        print(f"[DEBUG]   台账: {case.notice_number} (区号: {case.region})")
    for pdf_key, info in pdf_results.items():
        print(f"[DEBUG]   PDF: {pdf_key} -> 责令号: {info.notice_numbers}")
    
    # 收集匹配信息
    matched_rows = []  # 台账匹配到的
    pdf_matched_notices: Set[str] = set()  # 记录PDF中哪些责令号被匹配到了

    for case in registry.cases:
        matched_info, matched_notice = _match_case_to_pdf_with_notice(case, pdf_results)
        if matched_info:
            pdf_matched_notices.add(matched_notice)
            row = {
                '区号': case.region or '',
                '行政审查案号': matched_info.court_case_number or '',
                '责令号': case.notice_number or '',
                '被执行人': case.respondent or '',
                '职工姓名': case.employee or '',
                '金额': _format_amount(case.amount),
                '法官/法官助理': '',
                '执行时间': _format_date(matched_info.ruling_date) or '',
                '裁定结果': matched_info.ruling_result or '',
                '备注': '',
            }
            judge_parts = []
            if matched_info.judge:
                judge_parts.append(matched_info.judge)
            if matched_info.clerk:
                judge_parts.append(matched_info.clerk)
            row['法官/法官助理'] = '/'.join(judge_parts) if judge_parts else ''
            matched_rows.append(row)

    # 收集PDF独有的责令号
    pdf_only_rows = []
    for info in pdf_results.values():
        for notice in info.notice_numbers:
            norm_notice = _normalize_for_match(notice)
            # 检查这个责令号是否被匹配到了台账
            if norm_notice not in pdf_matched_notices:
                # 检查是否已经在pdf_only_rows中
                if not any(_normalize_for_match(r['责令号']) == norm_notice for r in pdf_only_rows):
                    row = {
                        '区号': '',
                        '行政审查案号': info.court_case_number or '',
                        '责令号': notice,
                        '被执行人': '',
                        '职工姓名': '',
                        '金额': '',
                        '法官/法官助理': '',
                        '执行时间': _format_date(info.ruling_date) or '',
                        '裁定结果': info.ruling_result or '',
                        '备注': 'PDF独有',
                    }
                    judge_parts = []
                    if info.judge:
                        judge_parts.append(info.judge)
                    if info.clerk:
                        judge_parts.append(info.clerk)
                    row['法官/法官助理'] = '/'.join(judge_parts) if judge_parts else ''
                    pdf_only_rows.append(row)

    # 排序
    _sort_rows_by_region(matched_rows)

    columns_order = ['区号', '行政审查案号', '责令号', '被执行人', '职工姓名', '金额', '法官/法官助理', '执行时间', '裁定结果', '备注']

    # 使用 openpyxl 直接写入，避免 pandas.ExcelWriter 在 PyInstaller 下的问题
    from openpyxl import Workbook
    
    wb = Workbook()
    
    # Sheet1: 匹配数据
    ws_matched = wb.active
    ws_matched.title = '匹配数据'
    _write_sheet_data(ws_matched, matched_rows, columns_order)
    _auto_fit_columns(ws_matched)
    _apply_header_style(ws_matched)
    
    # Sheet2: PDF独有
    ws_pdf_only = wb.create_sheet('PDF独有')
    _write_sheet_data(ws_pdf_only, pdf_only_rows, columns_order)
    _auto_fit_columns(ws_pdf_only)
    _apply_header_style(ws_pdf_only)
    
    wb.save(str(output_path))

    print(f"[DEBUG] Excel导出完成: {output_path}")
    print(f"[DEBUG]   - 匹配数据: {len(matched_rows)} 条")
    print(f"[DEBUG]   - PDF独有: {len(pdf_only_rows)} 条")
    for row in pdf_only_rows:
        print(f"[DEBUG]     PDF独有: {row['责令号']}")
    
    return output_path


def _match_case_to_pdf_with_notice(case, pdf_results: Dict[str, RulingInfo]) -> tuple[Optional[RulingInfo], str]:
    """
    匹配案件到PDF，返回匹配到的PDF信息和匹配到的具体责令号
    """
    for info in pdf_results.values():
        for ocr_notice in info.notice_numbers:
            norm_ocr = _normalize_for_match(ocr_notice)
            norm_excel = _normalize_for_match(case.notice_number)
            if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                return info, norm_ocr
    return None, ''


def _match_case_to_pdf(case, pdf_results: Dict[str, RulingInfo]) -> Optional[RulingInfo]:
    """兼容旧代码的匹配函数"""
    info, _ = _match_case_to_pdf_with_notice(case, pdf_results)
    return info


def _normalize_for_match(text: str) -> str:
    """标准化责令号用于匹配（与 product.normalize_notice_for_match 一致）"""
    return normalize_notice_for_match(text)


def _format_amount(amount) -> str:
    if amount is None:
        return ''
    return str(int(amount)) if amount == int(amount) else str(amount)


def _format_date(date_str: Optional[str]) -> str:
    if not date_str:
        return ''
    try:
        return chinese_date_to_arabic(date_str)
    except Exception:
        return date_str


def _sort_rows_by_region(rows: List[Dict]):
    region_order = {'萝岗': 1, '越秀': 2, '荔湾': 3, '天河': 4, '海珠': 5,
                    '白云': 6, '黄埔': 7, '番禺': 8, '花都': 9, '南沙': 10,
                    '增城': 11, '从化': 12}
    rows.sort(key=lambda r: (region_order.get(r.get('区号', ''), 99), r.get('责令号', '')))


def _write_sheet_data(ws, rows: List[Dict], columns: List[str]):
    """使用 openpyxl 直接写入数据到 sheet"""
    from openpyxl.styles import Font, Alignment
    
    # 写入表头
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(name='微软雅黑', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据行
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(columns, 1):
            value = row_data.get(col_name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value if value else '')
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal='left', vertical='center')


def _auto_fit_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = 0
                for ch in str(cell.value):
                    cell_len += 2 if '\u4e00' <= ch <= '\u9fff' else 1
                max_length = max(max_length, cell_len)
        adjusted = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)


def _apply_header_style(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def run_enforcement_extraction(
    input_dir: Path,
    excel_path: Path,
    output_dir: Path,
    use_ocr: bool = False,
) -> Dict[str, Any]:
    """
    运行强制执行组完整提取+导出流程

    流程：
    1. 加载台账（非诉表格.xlsx）
    2. 批量OCR识别裁定PDF（行审案号、责令号、法官/法官助理、责令作出时间）
    3. 按责令号匹配台账与OCR结果
    4. 导出合并后Excel（两个Sheet：匹配数据 + PDF独有）
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("强制执行组 - 裁定信息识别与导出")
    print("=" * 60)

    print("\n[1/4] 加载台账数据...")
    registry = load_enforcement_cases(excel_path)

    print("\n[2/4] 识别裁定PDF...")
    pdf_results = {}
    extractor = RulingPDFExtractor(use_ocr=use_ocr)
    pdf_files = sorted(input_dir.glob("*.pdf"))
    print(f"  发现 {len(pdf_files)} 个PDF文件")

    for pdf_file in pdf_files:
        print(f"  处理: {pdf_file.name}")
        try:
            info = extractor.extract_from_pdf(pdf_file)
            key = info.court_case_number if info.court_case_number else pdf_file.stem
            pdf_results[key] = info
            print(f"    案号: {info.court_case_number}")
            print(f"    责令号: {info.notice_numbers}")
            print(f"    法官: {info.judge}, 法官助理/书记员: {info.clerk}")
            print(f"    裁定日期: {info.ruling_date}")
            print(f"    裁定结果: {info.ruling_result}")
        except Exception as e:
            print(f"    [ERROR] 处理失败: {e}")

    print("\n[3/4] 匹配责令号并合并数据...")
    stats = _match_and_stats(registry, pdf_results)

    print("\n[4/4] 导出结果...")

    excel_output_path = output_dir / _paths_cfg.get('excel_output_filename', '执行组识别结果.xlsx')
    build_output_excel(registry, pdf_results, excel_output_path)

    json_path = output_dir / _paths_cfg.get('json_output_filename', 'enforcement_extracted.json')
    _export_json(pdf_results, stats, json_path)

    summary_path = output_dir / 'enforcement_summary.json'
    _export_summary(registry, pdf_results, summary_path)

    print("\n" + "=" * 60)
    print("处理统计")
    print("=" * 60)
    print(f"  PDF文件数: {stats['total_pdfs']}")
    print(f"  台账行数: {stats['total_excel_rows']}")
    print(f"  成功匹配: {stats['matched_rows']}")
    print(f"  未匹配: {stats['unmatched_rows']}")
    print(f"  撤回执行: {stats['withdraw_count']}")
    print(f"\n输出文件:")
    print(f"  - {excel_output_path}")
    print(f"  - {json_path}")
    print(f"  - {summary_path}")

    return {
        'stats': stats,
        'output_files': {
            'excel': str(excel_output_path),
            'json': str(json_path),
            'summary': str(summary_path),
        }
    }


def _match_and_stats(registry: EnforcementCaseRegistry, pdf_results: Dict[str, RulingInfo]) -> Dict[str, Any]:
    matched = set()
    withdraw_count = sum(1 for info in pdf_results.values() if info.is_withdraw)
    for case in registry.cases:
        for info in pdf_results.values():
            for ocr_notice in info.notice_numbers:
                norm_ocr = _normalize_for_match(ocr_notice)
                norm_excel = _normalize_for_match(case.notice_number)
                if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                    matched.add(case.notice_number)
                    break

    return {
        'total_pdfs': len(pdf_results),
        'total_excel_rows': len(registry.cases),
        'matched_rows': len(matched),
        'unmatched_rows': len(registry.cases) - len(matched),
        'withdraw_count': withdraw_count,
    }


def _export_json(pdf_results: Dict[str, RulingInfo], stats: Dict, output_path: Path):
    data = {
        'export_time': datetime.now().isoformat(),
        'stats': stats,
        'pdf_results': {k: v.to_dict() for k, v in pdf_results.items()},
    }
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON导出: {output_path}")


def _export_summary(registry: EnforcementCaseRegistry, pdf_results: Dict[str, RulingInfo], output_path: Path):
    summary = {
        'export_time': datetime.now().isoformat(),
        'excel_rows': len(registry.cases),
        'pdf_files': len(pdf_results),
        'matched': [],
        'unmatched_excel': [],
        'unmatched_pdfs': [],
    }

    matched_excel = set()
    for case in registry.cases:
        found = False
        for info in pdf_results.values():
            for ocr_notice in info.notice_numbers:
                norm_ocr = _normalize_for_match(ocr_notice)
                norm_excel = _normalize_for_match(case.notice_number)
                if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                    found = True
                    matched_excel.add(case.notice_number)
                    summary['matched'].append({
                        '责令号': case.notice_number,
                        '职工': case.employee,
                        '行审案号': info.court_case_number,
                        '法官': info.judge,
                        '日期': info.ruling_date,
                    })
                    break
            if found:
                break
        if not found:
            summary['unmatched_excel'].append({
                '责令号': case.notice_number,
                '被执行人': case.respondent,
                '职工': case.employee,
            })

    for key, info in pdf_results.items():
        if not any(m['行审案号'] == info.court_case_number for m in summary['matched']):
            summary['unmatched_pdfs'].append({
                '文件': key,
                '行审案号': info.court_case_number,
                '责令号': info.notice_numbers,
            })

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  汇总JSON导出: {output_path}")


if __name__ == "__main__":
    input_dir = Path("样本材料/强制组-自动化/提取信息")
    excel_path = Path("样本材料/强制组-自动化/提取信息/非诉表格.xlsx")
    output_dir = Path("output/enforcement")

    if input_dir.exists() and excel_path.exists():
        result = run_enforcement_extraction(input_dir, excel_path, output_dir, use_ocr=False)
        print("\n[OK] 处理完成!")
    else:
        print(f"ERROR: 输入路径不存在")
        print(f"  输入目录: {input_dir} (存在: {input_dir.exists()})")
        print(f"  Excel文件: {excel_path} (存在: {excel_path.exists()})")
