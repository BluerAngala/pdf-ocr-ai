#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
非诉组 PDF 导出模块

处理逻辑：
1. 责催证据文件：每个 PDF 就是一个独立案件，不切割，直接重命名
2. 申请书：OCR 检测"强制执行申请书"标题定位页边界，fallback 到固定页数
3. 授权书：固定 1 页/公司，按顺序切割
4. 所函：固定 1 页/公司，按顺序切割
"""

import json
import os
import queue
import re
import shutil
import sys
import threading
import time
from difflib import SequenceMatcher
from multiprocessing import Pool
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeoutError
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Iterable

try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from core.region_extractor import RegionExtractor, REGIONS
from contextlib import contextmanager

from pypdf import PdfReader, PdfWriter

from non_litigation.output_plan import build_expected_output_tree
from non_litigation.product import load_non_litigation_cases
from core.text_postprocessor import TextPostProcessor
from core.system_resource import detect_system_resources

from core.paths import ROOT, USER_DATA_DIR

from core.config_loader import load_config
from core.task_cancel import CancelledError
from core.smart_corrector import SmartCorrector, CorrectionCandidate
_cfg = load_config()

# 全局智能纠错器实例（延迟初始化）
_smart_corrector: Optional[SmartCorrector] = None

def _get_smart_corrector(ledger_cases: Optional[List[Dict]] = None) -> Optional[SmartCorrector]:
    """获取或创建智能纠错器"""
    global _smart_corrector
    if _smart_corrector is None and ledger_cases is not None:
        _smart_corrector = SmartCorrector(ledger_cases=ledger_cases, enable_learning=True)
    return _smart_corrector

try:
    from core.pdf_ocr_ultra import UltraFastOCR, OCRConfig, ImagePreprocessor
    HAS_OCR = True
except ImportError:
    HAS_OCR = False
    ImagePreprocessor = None
    _log("[WARN] pdf_ocr_ultra 导入失败，将使用 Mock OCR")

try:
    from core.task_state import TaskStateManager, Task
    from non_litigation.streaming import StreamingBatchProcessor
    HAS_STREAMING = True
except ImportError:
    HAS_STREAMING = False
    TaskStateManager = None
    Task = None
    StreamingBatchProcessor = None


SOURCE_MAPPING = _cfg.source_mapping
APPLICATION_BOUNDARY_KEYWORDS = _cfg.boundary_keywords.get('申请书', [])
APPLICATION_SECONDARY_EVIDENCE = _cfg.application_secondary_boundary_evidence
NON_LITIGATION_RESULT_DIRNAME = _cfg.result_dirname
NON_LITIGATION_TEMP_DIRNAME = _cfg.temp_dirname
NON_LITIGATION_INPUT_DIRNAME = _cfg.input_dirname
NOTICE_PATTERN = _cfg.notice_pattern
NON_LITIGATION_CORRECTIONS = _cfg.ocr_corrections
PAGES_PER_CASE = _cfg.pages_per_case

_audit_log: List[Dict] = []

_suppress_print = False
_PDF_TIMEOUT_SECONDS = 300


def _log(msg: str, quiet: bool = False):
    if not quiet and not _suppress_print:
        print(msg)


def _get_doc_regions(doc_type: str):
    return [REGIONS[name] for name in _cfg.ocr_doc_regions.get(doc_type, []) if name in REGIONS]


def _build_ocr_config() -> "OCRConfig":
    return OCRConfig(
        dpi=_cfg.ocr_dpi,
        max_image_size=_cfg.ocr_max_image_size,
        parallel_workers=_cfg.ocr_parallel_workers,
        small_pdf_page_threshold=_cfg.ocr_small_pdf_page_threshold,
    )


def _build_ocr_processors() -> Tuple["UltraFastOCR", RegionExtractor]:
    config = _build_ocr_config()
    from core.pdf_ocr_ultra import _ocr_engine as _global_engine
    skip = _global_engine is not None
    return UltraFastOCR(config, skip_warmup=skip, log_fn=_log), RegionExtractor(dpi=_cfg.ocr_region_dpi, poppler_path=config.poppler_path)


def _collect_region_texts(
    ocr: "UltraFastOCR",
    extractor: RegionExtractor,
    pdf_path: Path,
    page_num: int,
    doc_type: str,
    *,
    full_image=None,
    region_names: Optional[List[str]] = None,
) -> Tuple[str, List[Dict]]:
    requested_names = region_names or _cfg.ocr_doc_regions.get(doc_type, [])
    selected_names = [name for name in requested_names if name in REGIONS]
    regions = [REGIONS[name] for name in selected_names]
    if not regions:
        return "", []

    if full_image is None:
        images = extractor.extract_multiple_regions(pdf_path, page_num, regions)
    else:
        images = extractor.crop_regions_from_image(full_image, regions)
    pieces = []
    logs = []
    doc_ocr = _cfg.get_doc_ocr(doc_type)
    max_img_size = doc_ocr.max_image_size if doc_ocr else _cfg.ocr_region_max_image_size
    for region_name, region, image in zip(selected_names, regions, images):
        result = ocr.recognize_image_region(
            image,
            page_num=page_num,
            max_image_size=max_img_size,
            apply_enhancement=False,
            apply_sharpen=False,
            method=f"region:{region.name}",
            optimize_output=False,
        )
        text = result.text.strip()
        if text:
            pieces.append(text)
        logs.append({
            'region': region.name,
            'region_key': region_name,
            'method': result.method,
            'duration': result.duration,
            'text_length': len(text),
            'text': text,
        })

    return "\n".join(pieces), logs


def get_audit_log() -> List[Dict]:
    return _audit_log.copy()


def _log_audit(event: str, detail: Dict):
    _audit_log.append({'event': event, **detail})


_PREFETCH_SENTINEL = object()


def _prefetch_pages(region_extractor, pdf_path: Path, page_nums: List[int],
                    out_queue: queue.Queue, cancel_check=None):
    """后台线程：按序预提取页面图片放入队列，让 OCR 消费时不用等图片转换。"""
    for pn in page_nums:
        if cancel_check and cancel_check():
            out_queue.put(_PREFETCH_SENTINEL)
            return
        try:
            img = region_extractor.extract_full_page(pdf_path, pn)
            out_queue.put((pn, img), timeout=60)
        except Exception as e:
            out_queue.put((pn, e), timeout=60)
    out_queue.put(_PREFETCH_SENTINEL)


def _get_prefetched(out_queue: queue.Queue, timeout: int = 120):
    """从预取队列取一页；返回 (page_num, image) 或 (page_num, error) 或 None 表示结束。"""
    try:
        item = out_queue.get(timeout=timeout)
    except queue.Empty:
        return None
    if item is _PREFETCH_SENTINEL:
        return None
    return item


def _prefetch_pages_batch(region_extractor, pdf_path: Path, page_nums: List[int],
                          out_queue: queue.Queue, batch_size: int = 20,
                          cancel_check=None):
    """后台线程：批量预取页面，合并为少量 pdftoppm 调用以减少子进程开销。

    与 _prefetch_pages 的区别：不是逐页调用 extract_full_page，
    而是用 render_pages_batch 一批渲染多页，消除 39 次重复的
    pdftoppm 子进程启动和 PDF 解析。
    """
    for batch_start in range(0, len(page_nums), batch_size):
        if cancel_check and cancel_check():
            break
        batch = page_nums[batch_start:batch_start + batch_size]
        try:
            rendered = region_extractor.render_pages_batch(pdf_path, batch)
            for pn in batch:
                img = rendered.get(pn)
                if img is None:
                    out_queue.put((pn, RuntimeError(f"批量渲染未返回第 {pn} 页")), timeout=60)
                else:
                    out_queue.put((pn, img), timeout=60)
        except Exception as e:
            for pn in batch:
                out_queue.put((pn, e), timeout=60)
    out_queue.put(_PREFETCH_SENTINEL)


@contextmanager
def open_pdf_reader(pdf_path: Path):
    reader = PdfReader(str(pdf_path))
    try:
        yield reader
    finally:
        if hasattr(reader, 'stream') and reader.stream:
            reader.stream.close()


def apply_ocr_corrections(text: str, doc_type: str = None) -> str:
    corrections = _cfg.get_doc_corrections(doc_type) if doc_type else NON_LITIGATION_CORRECTIONS
    for wrong, correct in corrections:
        if wrong in text:
            text = text.replace(wrong, correct)
    return text


def normalize_notice_number(text: str) -> str:
    text = text.replace(' ', '')
    text = text.replace('(', '〔').replace(')', '〕')
    text = text.replace('（', '〔').replace('）', '〕')
    text = text.replace('[', '〔').replace(']', '〕')
    return text


_RELAXED_NOTICE_PATTERN = re.compile(
    r'([\u4e00-\u9fff]{1,2}公积金中心[\u4e00-\u9fff]{2,4}[责贵]字'
    r'[〔\[(［【]\d{4}[〕\)\]］】]\d+(?:-\d+)?号)'
)

_KNOWN_DISTRICTS = [
    '天河', '海珠', '荔湾', '越秀', '白云', '黄埔', '番禺', '南沙',
    '萝岗', '花都', '增城', '从化', '开发',
]

_DISTRICT_VARIANTS = {
    '番禹': '番禺', '香禹': '番禺', '步禹': '番禺',
    '萝岗': '萝岗', '罗岗': '萝岗',
    '海珠': '海珠', '海蛛': '海珠',
    '越秀': '越秀', '越矛': '越秀',
    '南沙': '南沙', '南沙': '南沙',
}

_PREFIX_VARIANTS = {
    '穗公积金中心', '稳公积金中心', '德公积金中心', '稍公积金中心',
}

_FUZZY_MATCH_THRESHOLD = 75.0

_NOTICE_STRUCT_PATTERN = re.compile(
    r'([\u4e00-\u9fff]{1,2}公积金中心)([\u4e00-\u9fff]{2,4})([责贵]字[〔\[(［【])(\d{4})([〕\)\]］】])(\d+(?:-\d+)?)(号)'
)


def _parse_notice_components(notice: str) -> Optional[Dict]:
    m = _NOTICE_STRUCT_PATTERN.match(notice)
    if not m:
        return None
    return {
        'prefix': m.group(1),
        'district': m.group(2),
        'prefix2': m.group(3),
        'year': m.group(4),
        'suffix_bracket': m.group(5),
        'number': m.group(6),
        'trailing': m.group(7),
    }


def _extract_notice_candidates_relaxed(text: str) -> List[str]:
    return _RELAXED_NOTICE_PATTERN.findall(text)


def _structural_correct_notice(raw: str) -> str:
    s = raw
    for wrong_prefix in _PREFIX_VARIANTS:
        if wrong_prefix in s:
            s = s.replace(wrong_prefix, '穗公积金中心', 1)
            break
    s = s.replace('贵字', '责字')
    for variant, correct in _DISTRICT_VARIANTS.items():
        if variant != correct and variant in s:
            s = s.replace(variant, correct)
    return s


def _match_notice_from_ocr_text(
    combined_text: str,
    post_processor: TextPostProcessor,
    notice_to_target: Dict[str, str],
    ledger_notices: List[str],
) -> Dict[str, Optional[str]]:
    """
    四级责令号匹配策略，确保准确性：

    Level 1 - 严格正则：直接匹配标准格式，精确对应台账（无需人工核查）
    Level 2 - 同根主号匹配：如 1234-1号 匹配 1234号（无需人工核查）
    Level 3 - 宽松正则 + 模糊匹配：容忍 OCR 误识（需要人工核查）
    Level 4 - 结构化纠错：不在台账中的责令号（需要人工核查）

    安全原则：
    - 模糊匹配和结构化纠错都标记为需要人工核查
    - 只有严格匹配和同根主号匹配才自动通过

    返回: {'source': str, 'notice_number': str|None, 'target_name': str|None, 'needs_review': bool}
    """
    # Level 1: 严格正则（经 apply_ocr_corrections 预纠错后）
    corrected = apply_ocr_corrections(combined_text, doc_type=None)
    normalized_text = post_processor.normalize_brackets(corrected)
    strict_numbers = post_processor.extract_decision_numbers(normalized_text)

    for dn in strict_numbers:
        n = normalize_notice_number(dn)
        if n in notice_to_target:
            return {'source': '严格匹配', 'notice_number': dn, 'target_name': notice_to_target[n], 'needs_review': False}

    # Level 2: 同根主号匹配（如 1234-1号 匹配 1234号）
    for dn in strict_numbers:
        detected_root = _get_notice_root_number(dn)
        if detected_root in notice_to_target:
            return {'source': '同根主号匹配', 'notice_number': dn, 'target_name': notice_to_target[detected_root], 'needs_review': False}
        # 反向匹配：OCR识别到主号，台账有子号
        for ledger_notice in ledger_notices:
            ledger_root = _get_notice_root_number(ledger_notice)
            if normalize_notice_number(dn) == ledger_root:
                n = normalize_notice_number(ledger_notice)
                return {'source': '同根主号匹配', 'notice_number': ledger_notice, 'target_name': notice_to_target.get(n), 'needs_review': False}

    # Level 3: 宽松正则 + 模糊匹配台账（需要人工核查）
    relaxed_candidates = _extract_notice_candidates_relaxed(combined_text)
    if not relaxed_candidates:
        relaxed_candidates = _extract_notice_candidates_relaxed(corrected)

    for raw_candidate in relaxed_candidates:
        structured = _structural_correct_notice(raw_candidate)
        ledger_match = fuzzy_match_notice_number(structured, ledger_notices)
        if ledger_match:
            n = normalize_notice_number(ledger_match)
            target = notice_to_target.get(n)
            return {'source': '模糊匹配', 'notice_number': ledger_match, 'target_name': target, 'needs_review': True}

    # Level 3.5: 用严格正则提取到的但不在台账中的，也尝试模糊匹配
    for dn in strict_numbers:
        ledger_match = fuzzy_match_notice_number(dn, ledger_notices)
        if ledger_match:
            n = normalize_notice_number(ledger_match)
            target = notice_to_target.get(n)
            return {'source': '模糊匹配', 'notice_number': ledger_match, 'target_name': target, 'needs_review': True}

    # Level 4: 结构化纠错（不在台账中的责令号，需要人工核查）
    # 优先用严格正则结果（已通过纠错），其次用宽松正则+结构纠错
    if strict_numbers:
        corrected = _structural_correct_notice(strict_numbers[0])
        return {'source': '结构纠错', 'notice_number': corrected, 'target_name': None, 'needs_review': True}

    for raw_candidate in relaxed_candidates:
        corrected_notice = _structural_correct_notice(raw_candidate)
        return {'source': '结构纠错', 'notice_number': corrected_notice, 'target_name': None, 'needs_review': True}

    return {'source': 'unknown', 'notice_number': None, 'target_name': None, 'needs_review': True}


def _resolve_district(raw_district: str) -> Optional[str]:
    for district in _KNOWN_DISTRICTS:
        try:
            from rapidfuzz import fuzz as rf_fuzz
            if rf_fuzz.ratio(raw_district, district, score_cutoff=70.0):
                return district
        except ImportError:
            if raw_district in _DISTRICT_VARIANTS:
                return _DISTRICT_VARIANTS[raw_district]
    return _DISTRICT_VARIANTS.get(raw_district)


def _structured_match_notice(
    candidate: str,
    ledger_notices: List[str],
) -> Optional[str]:
    """
    结构化责令号匹配：按组件（年+编号+区名）逐一比对。
    
    匹配逻辑：
    1. 解析候选和台账的责令号为结构组件
    2. 年份必须完全一致
    3. 数字编号必须完全一致
    4. 区名通过模糊匹配纠错（容忍番禹→番禺等 OCR 误识）
    5. 满足以上条件 → 采用台账原文（保证准确性）
    """
    corrected = _structural_correct_notice(candidate)
    cand_parts = _parse_notice_components(corrected)
    if not cand_parts:
        return None

    cand_district = _resolve_district(cand_parts['district'])
    cand_year = cand_parts['year']
    cand_number = cand_parts['number']

    for ledger in ledger_notices:
        ledger_parts = _parse_notice_components(ledger)
        if not ledger_parts:
            continue
        if ledger_parts['year'] != cand_year:
            continue
        if ledger_parts['number'] != cand_number:
            continue
        if cand_district and ledger_parts['district'] != cand_district:
            continue
        return ledger

    return None


def fuzzy_match_notice_number(
    candidate: str,
    ledger_notices: List[str],
    threshold: float = _FUZZY_MATCH_THRESHOLD,
) -> Optional[str]:
    match = _structured_match_notice(candidate, ledger_notices)
    if match:
        return match
    return None


def _get_notice_root_number(notice_number: str) -> str:
    normalized = normalize_notice_number(notice_number)
    return re.sub(r'(\d+)-\d+号$', r'\1号', normalized)


def _score_notice_candidate(page_num: int, text: str, post_processor: TextPostProcessor) -> Dict:
    corrected_text = apply_ocr_corrections(text, doc_type='责催')
    structured = post_processor.extract_notice_fields(corrected_text)
    decision_numbers = [normalize_notice_number(item) for item in structured.get('decision_numbers', [])]
    
    return {
        'page': page_num,
        'decision_numbers': decision_numbers,
        'has_notice': bool(decision_numbers),
        'company_name': structured.get('company_name'),
    }


def _select_notice_candidate(candidate_pages: List[Dict]) -> Dict:
    if not candidate_pages:
        return {}

    all_candidates = []
    for page_entry in candidate_pages:
        for number in page_entry.get('decision_numbers', []):
            normalized = normalize_notice_number(number)
            root_number = _get_notice_root_number(normalized)
            all_candidates.append({
                'page': page_entry.get('page'),
                'number': normalized,
                'root_number': root_number,
                'is_base_number': normalized == root_number,
                'company_name': page_entry.get('company_name'),
            })

    if not all_candidates:
        return {}

    base_candidates = [c for c in all_candidates if c['is_base_number']]
    best = base_candidates[0] if base_candidates else all_candidates[0]

    return {
        'selected_notice': best['number'],
        'selected_page': best['page'],
        'selected_root_notice': best['root_number'],
        'candidate_notices': all_candidates,
    }


def _compact_text(text: str) -> str:
    return text.replace('\n', '').replace(' ', '')


def _get_region_text(logs: List[Dict], region_key: str) -> str:
    """从 page_logs 中提取指定区域的 OCR 文本"""
    for entry in logs:
        if entry.get('region_key') == region_key:
            return entry.get('text', '')
    return ''


def _count_secondary_evidence(text: str, evidence_keywords: List[str]) -> int:
    """统计文本中命中的辅助证据关键词数量"""
    return sum(1 for kw in evidence_keywords if kw in text)


def _check_boundary_secondary(text: str, evidence_keywords: List[str], min_hits: int = 2) -> bool:
    """检查辅助证据是否足够判定为边界页"""
    if not evidence_keywords:
        return False
    return _count_secondary_evidence(text, evidence_keywords) >= min_hits


def _build_region_stats(page_logs: List[Dict]) -> Dict:
    text_lengths = [item.get('text_length', 0) for item in page_logs]
    return {
        'attempt_count': len(page_logs),
        'total_duration': round(sum(item.get('duration', 0) for item in page_logs), 4),
        'max_text_length': max(text_lengths, default=0),
        'total_text_length': sum(text_lengths),
        'regions': [item.get('region') for item in page_logs],
    }


def _should_fallback_notice(region_text: str, page_candidate: Optional[Dict], min_text_length: int = None) -> Tuple[bool, Optional[str]]:
    compact = _compact_text(region_text)
    
    if page_candidate and page_candidate.get('decision_numbers'):
        return False, None
    
    fallback_min = min_text_length or _cfg.notice_region_fallback_min_text_length
    if len(compact) < fallback_min:
        if '责字' in compact or '责令' in compact or '公积金' in compact or '穗' in compact or '越秀' in compact:
            return False, 'weak_notice_signal_only'
        return True, 'region_text_too_short'
    
    if '责字' in compact or '责令' in compact or '公积金' in compact:
        return False, 'weak_notice_signal_only'
    
    if '穗' in compact or '越秀' in compact or '中心' in compact:
        return False, 'weak_location_signal'
    
    return True, 'no_notice_signal'


def _should_fallback_application(
    page_num: int,
    text: str,
    detected_boundaries: List[int],
    expected_boundaries: int,
    expected_start_pages: set,
    min_text_length: int = None,
    page_logs: List[Dict] = None,
) -> Tuple[bool, Optional[str]]:
    compact = _compact_text(text)
    is_candidate_boundary_page = page_num in expected_start_pages
    previous_page_detected = (page_num - 1) in detected_boundaries
    next_page_detected = (page_num + 1) in detected_boundaries
    boundary_gap_exists = len(detected_boundaries) < expected_boundaries
    nearby_boundary_signal = previous_page_detected or next_page_detected
    fallback_min = min_text_length or _cfg.application_region_fallback_min_text_length
    weak_region_text = len(compact) < fallback_min

    if not is_candidate_boundary_page:
        return False, None
    # 主信号：标题区域关键词
    title_text = _get_region_text(page_logs or [], 'application_title')
    if title_text:
        title_corrected = apply_ocr_corrections(title_text, doc_type='申请书')
        if any(keyword in title_text or keyword in title_corrected for keyword in APPLICATION_BOUNDARY_KEYWORDS):
            return False, None
    elif any(keyword in text for keyword in APPLICATION_BOUNDARY_KEYWORDS):
        return False, None
    # 备用信号：被执行人区域辅助证据
    respondent_text = _get_region_text(page_logs or [], 'application_respondent')
    if respondent_text and APPLICATION_SECONDARY_EVIDENCE:
        respondent_corrected = apply_ocr_corrections(respondent_text, doc_type='申请书')
        combined_resp = respondent_text + respondent_corrected
        if _check_boundary_secondary(combined_resp, APPLICATION_SECONDARY_EVIDENCE, min_hits=2):
            return False, 'secondary_evidence_found'
    if weak_region_text:
        return True, 'boundary_candidate_text_short'
    if boundary_gap_exists and not nearby_boundary_signal:
        return True, 'boundary_gap_without_neighbor_signal'
    return False, None


def _should_fallback_company_doc(combined_text: str, region_usable: bool, marker_detected: bool) -> Tuple[bool, Optional[str]]:
    if marker_detected or region_usable:
        return False, None
    if len(_compact_text(combined_text)) < _cfg.company_doc_region_fallback_min_text_length:
        return True, 'region_text_too_short'
    return False, 'marker_missing_and_region_unusable'


def _build_ocr_output_summary(pdf_path: Path, pages: List[Dict], total_duration: float, *, doc_type: Optional[str] = None) -> Dict:
    if not pages:
        return {
            'doc_type': doc_type,
            'file_name': pdf_path.name,
            'page_count': 0,
            'total_duration': round(total_duration, 4),
            'fallback_pages': 0,
            'fallback_rate': 0.0,
            'region_attempts_total': 0,
            'region_ocr_duration_total': 0.0,
            'slowest_page': None,
        }

    fallback_pages = sum(1 for page in pages if page.get('fallback_used'))
    slowest_page = max(pages, key=lambda page: page.get('duration', 0))
    return {
        'doc_type': doc_type,
        'file_name': pdf_path.name,
        'page_count': len(pages),
        'total_duration': round(total_duration, 4),
        'fallback_pages': fallback_pages,
        'fallback_rate': round(fallback_pages / len(pages), 4),
        'region_attempts_total': sum(page.get('region_stats', {}).get('attempt_count', 0) for page in pages),
        'region_ocr_duration_total': round(sum(page.get('region_stats', {}).get('total_duration', 0) for page in pages), 4),
        'slowest_page': {
            'page': slowest_page.get('page'),
            'duration': round(slowest_page.get('duration', 0), 4),
            'method': slowest_page.get('method'),
            'fallback_used': slowest_page.get('fallback_used', False),
            'fallback_trigger_reason': slowest_page.get('fallback_trigger_reason'),
        },
    }


def fuzzy_match_notice(detected: str, target_map: Dict[str, str], threshold: float = 0.85) -> Tuple[Optional[str], float]:
    best_match = None
    best_ratio = 0
    detected_root = _get_notice_root_number(detected)

    for target in target_map.keys():
        if _get_notice_root_number(target) != detected_root:
            continue
        ratio = SequenceMatcher(None, detected, target).ratio()
        target_is_base = normalize_notice_number(target) == _get_notice_root_number(target)
        if not target_is_base and normalize_notice_number(detected) == detected_root:
            ratio -= 0.03
        if ratio > best_ratio and ratio >= threshold:
            best_ratio = ratio
            best_match = target_map[target]

    return (best_match, best_ratio) if best_match else (None, 0)


def _append_detected_notice_suffix(filename: str, detected_notice: str) -> str:
    """
    P0-#9: same_root_remap / fuzzy 时把 OCR 实际识别号附加到文件名后缀，
    避免「文件命名跟 PDF 实际不一致」混淆。
    """
    if not detected_notice:
        return filename
    safe = detected_notice.replace('/', '_').replace('\\', '_').strip()
    stem = Path(filename).stem
    suffix = Path(filename).suffix
    return f"{stem}[实际{safe}]{suffix}"


def _dedupe_destination(dst: Path) -> Path:
    """若目标已存在，自动加 -1 / -2 / ... 后缀找到一个未占用名；绝不覆盖既有产出。"""
    if not dst.exists():
        return dst
    parent = dst.parent
    stem = dst.stem
    suffix = dst.suffix
    n = 1
    while True:
        candidate = parent / f"{stem}-{n}{suffix}"
        if not candidate.exists():
            return candidate
        n += 1
        if n > 999:
            return parent / f"{stem}-{int(time.time())}{suffix}"


def _get_ocr_result(ocr_results: Dict[str, Dict], stem: str) -> Optional[Dict]:
    key = f'{stem}.pdf'
    if key in ocr_results:
        return ocr_results[key]
    bare = stem.replace('_ultra_result', '')
    if bare in ocr_results:
        return ocr_results[bare]
    if f'{bare}.pdf' in ocr_results:
        return ocr_results[f'{bare}.pdf']
    return None


def inspect_pdf_page_count(pdf_path: Path) -> int:
    with open_pdf_reader(pdf_path) as reader:
        return len(reader.pages)


def pre_check_input_materials(
    sample_root: Path,
    input_dir: Path,
    excel_path: Optional[Path] = None,
) -> List[str]:
    """处理前预检输入材料，检测案件数量是否匹配，返回警告信息列表"""
    warnings: List[str] = []
    
    try:
        cases = load_non_litigation_cases(sample_root, excel_path=excel_path)
        expected_count = len(cases)
    except Exception as e:
        warnings.append(f"无法加载台账案件: {e}")
        return warnings
    
    if expected_count == 0:
        warnings.append("台账中没有案件数据，请检查 Excel 文件是否正确")
        return warnings
    
    # 检查责催
    notice_dirs = get_notice_input_dirs(input_dir)
    notice_pdfs = []
    for notice_dir in notice_dirs:
        if notice_dir.exists():
            notice_pdfs.extend(list(notice_dir.glob('*.pdf')))
    
    if notice_pdfs:
        notice_count = len(notice_pdfs)
        # 责催每个 PDF 可能有多个案件，这里只统计文件数作为参考
        warnings.append(f"[预检] 责催: 发现 {notice_count} 个 PDF 文件，台账期望 {expected_count} 个案件")
    else:
        warnings.append("[预检] 责催: 未找到责催 PDF 文件")
    
    # 检查申请书
    app_source = input_dir / SOURCE_MAPPING.get('输出文件（申请书）', '申请书.pdf')
    if app_source.exists():
        app_pages = inspect_pdf_page_count(app_source)
        app_cases = app_pages // PAGES_PER_CASE.get('申请书', 2)
        status = '[OK]' if app_cases == expected_count else '[WARN]'
        warnings.append(
            f"[预检] 申请书: {app_pages} 页 ≈ {app_cases} 个案件，"
            f"台账期望 {expected_count} 个案件 {status}"
        )
    else:
        warnings.append("[预检] 申请书: 未找到申请书 PDF 文件")
    
    # 检查授权书
    auth_source = input_dir / SOURCE_MAPPING.get('输出文件（授权书）', '授权书.pdf')
    if auth_source.exists():
        auth_pages = inspect_pdf_page_count(auth_source)
        status = '[OK]' if auth_pages == expected_count else '[WARN]'
        warnings.append(
            f"[预检] 授权书: {auth_pages} 页，台账期望 {expected_count} 个案件 {status}"
        )
    else:
        warnings.append("[预检] 授权书: 未找到授权书 PDF 文件")
    
    # 检查所函
    letter_source = input_dir / SOURCE_MAPPING.get('输出文件（所函）', '所函.pdf')
    if letter_source.exists():
        letter_pages = inspect_pdf_page_count(letter_source)
        status = '[OK]' if letter_pages == expected_count else '[WARN]'
        warnings.append(
            f"[预检] 所函: {letter_pages} 页，台账期望 {expected_count} 个案件 {status}"
        )
    else:
        warnings.append("[预检] 所函: 未找到所函 PDF 文件")
    
    return warnings


def get_non_litigation_input_root(project_root: Path) -> Path:
    return project_root / 'input' / NON_LITIGATION_INPUT_DIRNAME


def get_non_litigation_result_root(project_root: Path) -> Path:
    return USER_DATA_DIR / 'output' / NON_LITIGATION_RESULT_DIRNAME


def get_non_litigation_temp_root(project_root: Path) -> Path:
    return USER_DATA_DIR / 'temp' / NON_LITIGATION_TEMP_DIRNAME


def ensure_non_litigation_input_structure(project_root: Path) -> Path:
    input_root = get_non_litigation_input_root(project_root)
    input_root.mkdir(parents=True, exist_ok=True)
    for item in input_root.parent.iterdir():
        if item.is_file() and item.suffix.lower() == '.pdf' and not (input_root / item.name).exists():
            shutil.move(str(item), str(input_root / item.name))
    return input_root


def get_notice_input_dirs(input_dir: Path) -> List[Path]:
    candidate_dirs = [input_dir]
    for subdir_name in ['责催（证据材料）', 'notice-evidence']:
        nested_dir = input_dir / subdir_name
        if nested_dir.exists() and nested_dir.is_dir():
            candidate_dirs.append(nested_dir)
    return candidate_dirs


def iter_notice_pdf_paths(input_dir: Path) -> Iterable[Path]:
    seen: set[str] = set()
    for notice_dir in get_notice_input_dirs(input_dir):
        for pdf_path in sorted(notice_dir.glob('*.pdf')):
            if pdf_path.name in SOURCE_MAPPING.values() or pdf_path.name in seen:
                continue
            seen.add(pdf_path.name)
            yield pdf_path


def normalize_company_name_for_matching(value: str) -> str:
    text = str(value).strip()
    text = text.replace('\n', '').replace('\r', '').replace(' ', '')
    text = text.replace('（', '(').replace('）', ')')
    return text


def detect_page_ranges(total_pages: int, expected_count: int, doc_type: str) -> List[Tuple[int, int]]:
    pages_per_item = PAGES_PER_CASE.get(doc_type, 1)
    expected_pages = expected_count * pages_per_item

    if total_pages != expected_pages:
        _log(
            f'  [INFO] {doc_type}: 实际 {total_pages} 页，台账期望 {expected_pages} 页 '
            f'({expected_count} 个 × {pages_per_item} 页)，按实际页数处理'
        )

    actual_count = min(expected_count, total_pages // pages_per_item) if total_pages > 0 else 0
    ranges = []
    for i in range(actual_count):
        start = i * pages_per_item
        end = start + pages_per_item
        ranges.append((start, end))
    return ranges


def detect_application_page_ranges_by_ocr(ocr_results: Dict[str, Dict], total_pages: int, expected_cases: int) -> List[Tuple[int, int]]:
    """
    通过 OCR 识别"强制执行申请书"标题来定位页边界。
    如果 OCR 结果不可用或检测到的边界数不匹配，fallback 到固定页数。
    """
    data = _get_ocr_result(ocr_results, '申请书')
    if not data or not data.get('pages'):
        _log(f"  [WARN] 无申请书 OCR 缓存，使用固定 {PAGES_PER_CASE['申请书']} 页/案件")
        return detect_page_ranges(total_pages, expected_cases, '申请书')

    boundary_pages = []
    for page_data in data['pages']:
        text = page_data.get('text', '')
        page_num = page_data.get('page', 0)
        for keyword in APPLICATION_BOUNDARY_KEYWORDS:
            if keyword in text:
                boundary_pages.append(page_num - 1)
                break

    if len(boundary_pages) == expected_cases:
        ranges = []
        for i, start in enumerate(boundary_pages):
            end = boundary_pages[i + 1] if i + 1 < len(boundary_pages) else total_pages
            ranges.append((start, end))
        _log(f"  [OK] OCR 检测到 {len(boundary_pages)} 个申请书边界")
        return ranges

    if boundary_pages:
        _log(f"  [INFO] OCR 检测到 {len(boundary_pages)} 个边界，台账期望 {expected_cases} 个，按实际处理")
    else:
        _log(f"  [INFO] OCR 未检测到申请书边界，使用固定 {PAGES_PER_CASE['申请书']} 页/案件")

    return detect_page_ranges(total_pages, expected_cases, '申请书')


def discover_notice_files(input_dir: Path) -> List[str]:
    """
    动态发现输入目录中的责催 PDF 文件。
    按文件名自然排序（1.pdf, 2.pdf, ..., 10.pdf, ...）
    """
    def natural_sort_key(name: str) -> List:
        parts = []
        for part in re.split(r'(\d+)', name):
            if part.isdigit():
                parts.append(int(part))
            else:
                parts.append(part)
        return parts

    pdf_files = sorted(
        [pdf_path.name for pdf_path in iter_notice_pdf_paths(input_dir)],
        key=natural_sort_key,
    )
    return pdf_files


def detect_notice_source_mapping_from_ocr(
    ocr_results: Dict[str, Dict],
    notice_files: List[str],
    ledger_notices: Optional[List[str]] = None,
    notice_to_target: Optional[Dict[str, str]] = None,
    ledger_cases: Optional[List[Dict]] = None,
) -> Dict[str, str]:
    """
    从 OCR 结果中识别责催文件的责令号，并与台账进行比对纠错

    Args:
        ocr_results: OCR 识别结果
        notice_files: 责催 PDF 文件列表
        ledger_notices: 台账中的责令号列表（用于比对纠错）
        notice_to_target: 责令号到目标文件名的映射

    Returns:
        {source_filename: detected_notice_number}
    """
    mapping: Dict[str, str] = {}
    post_processor = TextPostProcessor()

    # 构建台账查找集合（用于快速判断）
    ledger_set = set(normalize_notice_number(n) for n in (ledger_notices or []))

    for source_name in notice_files:
        stem = source_name.replace('.pdf', '')
        data = _get_ocr_result(ocr_results, stem)

        if not data:
            continue

        candidate_pages = []
        for page in data['pages']:
            text = page.get('text', '').replace('\n', ' ')
            candidate = _score_notice_candidate(page.get('page', 0), text, post_processor)
            if candidate.get('decision_numbers'):
                candidate_pages.append(candidate)

        selection = _select_notice_candidate(candidate_pages)
        selected_notice = selection.get('selected_notice') or data.get('selected_notice')
        selected_page = selection.get('selected_page') or data.get('selected_page')

        if selected_notice:
            # 如果提供了台账数据，进行比对纠错
            if ledger_notices:
                final_match = _match_notice_with_ledger(
                    selected_notice,
                    ledger_notices,
                    notice_to_target or {},
                    source_name,
                    ledger_cases,
                )
                matched_notice = final_match.get('notice_number') or selected_notice
                match_source = final_match.get('source', 'ocr_only')
                needs_review = final_match.get('needs_review', False)
                review_reason = final_match.get('review_reason')

                # 如果匹配结果与原始识别不同，记录审计日志
                if matched_notice != selected_notice:
                    _log_audit('notice_corrected', {
                        'source': source_name,
                        'original_detected': selected_notice,
                        'corrected_to': matched_notice,
                        'match_source': match_source,
                        'needs_review': needs_review,
                        'review_reason': review_reason,
                    })
                    if needs_review:
                        _log(f"  [REVIEW] {source_name}: '{selected_notice}' -> '{matched_notice}' ({match_source}) - 需人工核查: {review_reason}")
                    else:
                        _log(f"  [CORRECTED] {source_name}: '{selected_notice}' -> '{matched_notice}' ({match_source})")

                mapping[source_name] = matched_notice

                if selected_page:
                    _log(f"    [INFO] 采用第 {selected_page} 页候选")
            else:
                # 无台账数据，直接使用 OCR 结果
                mapping[source_name] = selected_notice
                _log(f"  [OK] {source_name}: 识别到责令号 '{selected_notice}'")
                if selected_page:
                    _log(f"    [INFO] 采用第 {selected_page} 页候选")
        else:
            _log(f"  [WARN] {source_name}: 未识别到责令号")

    return mapping


def _match_notice_with_ledger(
    detected_notice: str,
    ledger_notices: List[str],
    notice_to_target: Dict[str, str],
    source_name: str,
    ledger_cases: Optional[List[Dict]] = None,
) -> Dict[str, Optional[str]]:
    """
    将 OCR 识别的责令号与台账进行比对，返回最佳匹配结果

    匹配策略（按优先级）：
    1. 精确匹配：标准化后的责令号与台账完全一致
    2. 智能纠错（L1）：高置信度自动纠错（区名纠错等）
    3. 同根主号匹配：如 1234-1号 匹配 1234号
    4. 结构化匹配：按年+编号+区名组件比对
    5. 智能纠错（L2/L3）：中低置信度，建议人工核查
    6. 模糊匹配：相似度 >= 85%

    安全原则：
    - L1（高置信度）：自动应用
    - L2/L3（中低置信度）：返回匹配结果但标记需要人工核查

    Returns:
        {
            'source': str,
            'notice_number': str|None,
            'correction_info': dict|None,
            'needs_review': bool,  # 是否需要人工核查
            'review_reason': str|None  # 人工核查原因
        }
    """
    normalized_detected = normalize_notice_number(detected_notice)

    # Level 1: 精确匹配
    for ledger in ledger_notices:
        if normalize_notice_number(ledger) == normalized_detected:
            return {
                'source': 'exact_match',
                'notice_number': ledger,
                'correction_info': None,
                'needs_review': False,
                'review_reason': None
            }

    # Level 2: 智能纠错（L1 - 高置信度自动应用）
    if ledger_cases:
        corrector = _get_smart_corrector(ledger_cases)
        if corrector:
            correction = corrector.correct_notice_number(detected_notice)

            # L1级别：高置信度，自动应用
            if correction.level == 'L1' and corrector.should_auto_apply(correction):
                corrected_normalized = normalize_notice_number(correction.corrected)
                for ledger in ledger_notices:
                    if normalize_notice_number(ledger) == corrected_normalized:
                        return {
                            'source': f'smart_L1_{correction.method}',
                            'notice_number': ledger,
                            'correction_info': {
                                'original': correction.original,
                                'corrected': correction.corrected,
                                'confidence': correction.confidence,
                                'reason': correction.reason,
                                'level': correction.level,
                            },
                            'needs_review': False,
                            'review_reason': None
                        }

            # L2/L3级别：需要人工核查
            if correction.level in ('L2', 'L3'):
                corrected_normalized = normalize_notice_number(correction.corrected)
                for ledger in ledger_notices:
                    if normalize_notice_number(ledger) == corrected_normalized:
                        return {
                            'source': f'smart_{correction.level}_{correction.method}',
                            'notice_number': ledger,
                            'correction_info': {
                                'original': correction.original,
                                'corrected': correction.corrected,
                                'confidence': correction.confidence,
                                'reason': correction.reason,
                                'level': correction.level,
                            },
                            'needs_review': True,
                            'review_reason': f'智能纠错{correction.level}级别：{correction.reason}'
                        }

    # Level 3: 同根主号匹配
    detected_root = _get_notice_root_number(detected_notice)
    for ledger in ledger_notices:
        ledger_root = _get_notice_root_number(ledger)
        if detected_root == ledger_root:
            return {
                'source': 'root_match',
                'notice_number': ledger,
                'correction_info': None,
                'needs_review': False,
                'review_reason': None
            }

    # Level 4: 结构化匹配（年+编号+区名）
    structured_match = _structured_match_notice(detected_notice, ledger_notices)
    if structured_match:
        return {
            'source': 'structured_match',
            'notice_number': structured_match,
            'correction_info': None,
            'needs_review': False,
            'review_reason': None
        }

    # Level 5: 模糊匹配（相似度 >= 85%）
    best_match = None
    best_ratio = 0.0
    for ledger in ledger_notices:
        ratio = SequenceMatcher(None, normalized_detected, normalize_notice_number(ledger)).ratio()
        if ratio > best_ratio and ratio >= 0.85:  # 85% 阈值
            best_ratio = ratio
            best_match = ledger

    if best_match:
        return {
            'source': f'fuzzy_match({best_ratio:.1%})',
            'notice_number': best_match,
            'correction_info': None,
            'needs_review': True,  # 模糊匹配需要人工核查
            'review_reason': f'模糊匹配，相似度{best_ratio:.1%}'
        }

    # 无匹配，返回原始识别结果
    return {
        'source': 'no_match',
        'notice_number': detected_notice,
        'correction_info': None,
        'needs_review': True,  # 未匹配到台账，需要人工核查
        'review_reason': '未匹配到台账记录'
    }


def build_mock_ocr_results(sample_root: Path, input_dir: Path | None = None) -> Dict[str, Dict]:
    """构建 Mock OCR 结果（用于测试）"""
    standard_root = sample_root / _cfg.standard_output_dirname
    ocr_results: Dict[str, Dict] = {}

    ocr_noise_samples = _cfg.mock_noise_samples

    def get_subdir(doc_type: str) -> str:
        return _cfg.standard_output_subdirs.get(doc_type, _cfg.directory_mapping.get(doc_type, doc_type))

    application_pages = []
    for index, pdf_path in enumerate(sorted((standard_root / get_subdir('申请书')).glob('*.pdf'))):
        page_count = inspect_pdf_page_count(pdf_path)
        for page_offset in range(page_count):
            page_number = len(application_pages) + 1
            if page_offset == 0:
                noise_idx = index % len(ocr_noise_samples)
                text = f'强制执行申请书\n名称：案子{index + 1}\n穗公积金中心{ocr_noise_samples[noise_idx]}越秀责字'
            else:
                text = f'被执行人：公司{index + 1}\n金额：10000元'
            application_pages.append({'page': page_number, 'text': text})
    ocr_results['申请书.pdf'] = {'pages': application_pages, 'total_pages': len(application_pages), 'filename': '申请书.pdf'}

    for doc_type_key in ['授权书', '所函']:
        marker = _cfg.doc_type_map[doc_type_key].content_marker
        folder = get_subdir(doc_type_key)
        pages = []
        for index, pdf_path in enumerate(sorted((standard_root / folder).glob('*.pdf'))):
            company_name = pdf_path.stem
            noise_idx = index % len(ocr_noise_samples)
            text = f'{marker}\n{ocr_noise_samples[noise_idx]}\n{company_name}'
            pages.append({'page': index + 1, 'text': text})
        ocr_results[f'{doc_type_key}.pdf'] = {'pages': pages, 'total_pages': len(pages), 'filename': f'{doc_type_key}.pdf'}

    notice_files = sorted((standard_root / get_subdir('责催')).glob('*.pdf'))

    if input_dir and input_dir.exists():
        std_page_map: Dict[Path, int] = {}
        std_used: Dict[int, Path] = {}
        for pdf_path in notice_files:
            pc = inspect_pdf_page_count(pdf_path)
            std_page_map[pdf_path] = pc

        src_files = discover_notice_files(input_dir)
        for src_name in src_files:
            src_path = input_dir / src_name
            if not src_path.exists():
                continue
            src_pages = inspect_pdf_page_count(src_path)
            matched_std = None
            for std_path, std_pages in std_page_map.items():
                if std_pages == src_pages and std_pages not in std_used:
                    matched_std = std_path
                    std_used[std_pages] = std_path
                    break

            if matched_std:
                stem_name = matched_std.stem
                if '-责催-' in stem_name:
                    notice_number = stem_name.split('-责催-')[1]
                    normalized_number = normalize_notice_number(notice_number)
                    ocr_results[src_name] = {
                        'pages': [{'page': 1, 'text': normalized_number}],
                        'total_pages': 1,
                        'filename': src_name,
                    }
            else:
                _log(f"  [WARN] Mock: 无法按页数匹配 {src_name}，使用顺序匹配")
                fallback_idx = src_files.index(src_name)
                if fallback_idx < len(notice_files):
                    notice_number = notice_files[fallback_idx].stem.split('-责催-')[1]
                    normalized_number = normalize_notice_number(notice_number)
                    ocr_results[src_name] = {
                        'pages': [{'page': 1, 'text': normalized_number}],
                        'total_pages': 1,
                        'filename': src_name,
                    }
    else:
        notice_numbers = [pdf_path.stem.split('-责催-')[1] for pdf_path in notice_files]
        for idx, notice_number in enumerate(notice_numbers):
            normalized_number = normalize_notice_number(notice_number)
            src_name = f'{idx + 1}.pdf'
            ocr_results[src_name] = {
                'pages': [{'page': 1, 'text': normalized_number}],
                'total_pages': 1,
                'filename': src_name,
            }

    return ocr_results


def run_real_ocr_on_pdf(pdf_path: Path, use_mock: bool = False,
                        is_notice: bool = False, stop_pattern: Optional[re.Pattern] = None,
                        doc_type: Optional[str] = None, ocr: Optional["UltraFastOCR"] = None,
                        region_extractor: Optional[RegionExtractor] = None,
                        post_processor: Optional[TextPostProcessor] = None,
                        quiet: bool = False,
                        page_progress: Optional[callable] = None,
                        cancel_check: Optional[callable] = None) -> Dict:
    if use_mock or not HAS_OCR:
        return {'pages': [], 'total_pages': 0, 'filename': pdf_path.name, 'method': 'mock'}

    _log(f"  [OCR] 开始识别: {pdf_path.name}")

    try:
        if ocr is None or region_extractor is None:
            shared_ocr, shared_region_extractor = _build_ocr_processors()
            ocr = ocr or shared_ocr
            region_extractor = region_extractor or shared_region_extractor

        if is_notice and stop_pattern:
            _log(f"  [OCR] 使用逐页识别模式（短窗口扫描后停止）")
            notice_post_processor = post_processor or TextPostProcessor()

            doc_ocr = _cfg.get_doc_ocr(doc_type or '责催')
            if doc_ocr.enable_region_first:
                pages = []
                total_start = time.perf_counter()
                page_num = 1
                stopped_early = False
                candidate_pages = []
                first_hit_page = None
                stop_after_page = None
                max_scan_pages = max(1, doc_ocr.scan_max_pages)
                scan_window_pages = max(0, doc_ocr.scan_window_pages)

                while page_num <= max_scan_pages:
                    if cancel_check and cancel_check():
                        raise CancelledError("用户取消")
                    try:
                        full_image = region_extractor.extract_full_page(pdf_path, page_num)
                    except Exception:
                        break

                    if doc_ocr.skip_blank_pages:
                        if ImagePreprocessor.is_blank_page(full_image, threshold=doc_ocr.blank_page_threshold):
                            pages.append({
                                'page': page_num,
                                'text': '',
                                'method': 'blank_page_skipped',
                                'duration': 0,
                                'region_attempts': [],
                                'region_stats': {'attempt_count': 0, 'total_duration': 0, 'max_text_length': 0, 'total_text_length': 0, 'regions': []},
                                'region_text_length': 0,
                                'fallback_used': False,
                                'fallback_trigger_reason': None,
                                'notice_matched': False,
                                'notice_candidates': [],
                                'notice_candidate_score': None,
                                'notice_page_profile': {},
                            })
                            page_num += 1
                            continue

                    page_logs = []
                    region_text = ''
                    if doc_type:
                        region_text, page_logs = _collect_region_texts(
                            ocr,
                            region_extractor,
                            pdf_path,
                            page_num,
                            doc_type,
                            full_image=full_image,
                        )
                    region_text = apply_ocr_corrections(region_text, doc_type=doc_type)

                    duration = sum(item['duration'] for item in page_logs)
                    text = region_text
                    method = 'region_first'
                    page_candidate = _score_notice_candidate(page_num, text, notice_post_processor) if text else None
                    matched = bool(page_candidate and page_candidate.get('decision_numbers'))
                    needs_fallback, fallback_trigger_reason = _should_fallback_notice(region_text, page_candidate, min_text_length=doc_ocr.region_fallback_min_text_length)

                    if needs_fallback and doc_ocr.allow_full_page_fallback:
                        full_result = ocr.recognize_full_page_image(
                            full_image,
                            page_num=page_num,
                            method='full_page_fallback',
                            optimize_output=True,
                        )
                        full_text = apply_ocr_corrections(full_result.text, doc_type=doc_type)
                        duration += full_result.duration
                        if full_text:
                            text = full_text
                            method = full_result.method
                            page_candidate = _score_notice_candidate(page_num, text, notice_post_processor)
                            matched = bool(page_candidate.get('decision_numbers'))

                    if matched and page_candidate:
                        candidate_pages.append(page_candidate)
                        pages.append({
                            'page': page_num,
                            'text': text,
                            'method': method,
                            'duration': duration,
                            'region_attempts': page_logs,
                            'region_stats': _build_region_stats(page_logs),
                            'region_text_length': len(_compact_text(region_text)),
                            'fallback_used': method == 'full_page_fallback',
                            'fallback_trigger_reason': fallback_trigger_reason if method == 'full_page_fallback' else None,
                            'notice_matched': matched,
                            'notice_candidates': page_candidate.get('decision_numbers', []) if page_candidate else [],
                        })
                        _log(f"    [OK] 第 {page_num} 页找到责令号，立即停止")
                        stopped_early = True
                        break

                    pages.append({
                        'page': page_num,
                        'text': text,
                        'method': method,
                        'duration': duration,
                        'region_attempts': page_logs,
                        'region_stats': _build_region_stats(page_logs),
                        'region_text_length': len(_compact_text(region_text)),
                        'fallback_used': method == 'full_page_fallback',
                        'fallback_trigger_reason': fallback_trigger_reason if method == 'full_page_fallback' else None,
                        'notice_matched': matched,
                        'notice_candidates': page_candidate.get('decision_numbers', []) if page_candidate else [],
                    })

                    page_num += 1

                total_duration = time.perf_counter() - total_start
                selection = _select_notice_candidate(candidate_pages)
                result = {
                    'filename': pdf_path.name,
                    'filepath': str(pdf_path),
                    'total_pages': len(pages),
                    'method': 'region_first_sequential',
                    'total_duration': total_duration,
                    'pages': pages,
                    'full_text': "\n\n".join([f"=== 第{p['page']}页 ===\n{p['text']}" for p in pages if p['text']]),
                    'stopped_early': stopped_early,
                    'selected_notice': selection.get('selected_notice'),
                    'selected_page': selection.get('selected_page'),
                    'selected_root_notice': selection.get('selected_root_notice'),
                    'candidate_notices': selection.get('candidate_notices', []),
                    'performance_summary': _build_ocr_output_summary(pdf_path, pages, total_duration, doc_type=doc_type),
                }
            else:
                result = ocr.process_pdf_pages_sequential(
                    str(pdf_path),
                    stop_condition=lambda page_num, text: bool(stop_pattern.search(apply_ocr_corrections(text, doc_type='责催'))),
                    max_pages=max(1, _cfg.notice_scan_max_pages),
                )
        elif doc_type in {'申请书', '授权书', '所函'} and _cfg.ocr_enable_region_first:
            total_pages = inspect_pdf_page_count(pdf_path)
            pages = []
            total_start = time.perf_counter()
            fallback_used = False
            application_region_results = []

            def _is_application_boundary(text: str, page_logs: List[Dict] = None) -> Tuple[bool, Optional[str]]:
                """
                两级边界检测：
                1. 标题区域匹配边界关键词 → 'primary'
                2. 被执行人区域辅助证据 ≥ 2 条 → 'secondary'
                3. 组合文本回退匹配 → 'combined'
                """
                # Level 1: 标题区域关键词（主信号）
                title_text = _get_region_text(page_logs or [], 'application_title')
                if title_text:
                    title_corrected = apply_ocr_corrections(title_text, doc_type='申请书')
                    for kw in APPLICATION_BOUNDARY_KEYWORDS:
                        if kw in title_text or kw in title_corrected:
                            return True, 'primary'

                # Level 2: 被执行人区域辅助证据（备用信号）
                respondent_text = _get_region_text(page_logs or [], 'application_respondent')
                if respondent_text and APPLICATION_SECONDARY_EVIDENCE:
                    respondent_corrected = apply_ocr_corrections(respondent_text, doc_type='申请书')
                    combined_resp = respondent_text + respondent_corrected
                    if _check_boundary_secondary(combined_resp, APPLICATION_SECONDARY_EVIDENCE, min_hits=2):
                        return True, 'secondary'

                # Level 3: 组合文本回退（兼容 quick_scan / 全页 fallback）
                if any(keyword in text for keyword in APPLICATION_BOUNDARY_KEYWORDS):
                    return True, 'combined'

                return False, None

            def _has_title_fragments(compact: str, fragments: List[str], *, min_hits: int) -> bool:
                return sum(1 for fragment in fragments if fragment in compact) >= min_hits

            def _check_region_usable(doc_type_key: str, logs: List[Dict], combined_text: str) -> bool:
                dt = _cfg.doc_type_map.get(doc_type_key)
                if not dt or not dt.ocr.usability_check:
                    return len(combined_text.replace('\n', '').replace(' ', '')) >= 6
                uc = dt.ocr.usability_check
                compact = combined_text.replace('\n', '').replace(' ', '')
                for kw in uc.keywords:
                    if kw in compact:
                        return True
                if uc.fragment_keywords and _has_title_fragments(compact, uc.fragment_keywords, min_hits=uc.min_fragment_hits):
                    return True
                if logs and logs[0].get('text_length', 0) >= uc.min_text_length:
                    return True
                return len(compact) >= uc.min_text_length

            if doc_type == '申请书':
                pages_per = PAGES_PER_CASE['申请书']
                expected_cases = max(1, total_pages // pages_per)
                _log(f"  [申请书] {total_pages} pages, {pages_per} pages/case, {expected_cases} cases expected")

                all_page_nums_app = list(range(1, min(total_pages, expected_cases * pages_per) + 1))
                # 大文件使用更大的批次和队列
                app_batch_size = 20 if total_pages > 50 else 10
                pf_queue_app: queue.Queue = queue.Queue(maxsize=max(8, app_batch_size))
                pf_thread_app = threading.Thread(
                    target=_prefetch_pages_batch,
                    args=(region_extractor, pdf_path, all_page_nums_app, pf_queue_app, app_batch_size),
                    kwargs={'cancel_check': cancel_check},
                    daemon=True,
                )
                pf_thread_app.start()
                prefetched_app: Dict[int, Any] = {}

                for case_idx in range(expected_cases):
                    if cancel_check and cancel_check():
                        _log(f"  [申请书] 已取消，已完成 {case_idx}/{expected_cases}")
                        raise CancelledError("用户取消")
                    start_page = 1 + case_idx * pages_per
                    if page_progress and case_idx % 10 == 0:
                        page_progress(start_page, total_pages)
                    for offset in range(pages_per):
                        page_num = start_page + offset
                        if cancel_check and cancel_check():
                            _log(f"  [申请书] 已取消，第 {page_num} 页")
                            raise CancelledError("用户取消")
                        if page_num > total_pages:
                            break
                        if any(p['page'] == page_num for p in pages):
                            continue
                        if page_num not in prefetched_app:
                            item = _get_prefetched(pf_queue_app, timeout=180)  # 3分钟超时
                            if item is not None:
                                pn, img_or_err = item
                                if not isinstance(img_or_err, Exception):
                                    prefetched_app[pn] = img_or_err
                        full_image = prefetched_app.pop(page_num, None)
                        if full_image is None:
                            full_image = region_extractor.extract_full_page(pdf_path, page_num)
                        is_group_start = (offset == 0)
                        if is_group_start:
                            region_text, page_logs = _collect_region_texts(
                                ocr,
                                region_extractor,
                                pdf_path,
                                page_num,
                                doc_type,
                                full_image=full_image,
                            )
                        else:
                            # 偶数页：OCR 执行依据页的责令号区域（用于后续匹配台账）
                            notice_region = REGIONS.get('application_execution_notice')
                            if notice_region:
                                notice_img = region_extractor.crop_region_from_image(full_image, notice_region)
                                result = ocr.recognize_image_region(
                                    notice_img, page_num,
                                    method='notice_extract', optimize_output=False,
                                )
                                region_text = result.text
                                page_logs = [{
                                    'region': notice_region.name,
                                    'region_key': 'application_execution_notice',
                                    'method': result.method,
                                    'duration': result.duration,
                                    'text_length': len(region_text),
                                    'text': region_text,
                                }]
                            else:
                                qs_cfg = _cfg.get_doc_ocr(doc_type).quick_scan if _cfg.get_doc_ocr(doc_type) else None
                                qs_size = tuple(qs_cfg.target_size) if qs_cfg and qs_cfg.enabled else (400, 400)
                                img = ImagePreprocessor.optimize_for_ocr(full_image, target_size=qs_size)
                                result = ocr.recognize_image_region(img, page_num, method='quick_scan', optimize_output=False)
                                region_text = result.text
                                page_logs = [{'region': 'quick_scan', 'text_length': len(region_text), 'duration': result.duration}]
                        corrected_text = apply_ocr_corrections(region_text, doc_type=doc_type)
                        duration = sum(item['duration'] for item in page_logs)
                        boundary = False
                        boundary_evidence = None
                        if is_group_start:
                            boundary, boundary_evidence = _is_application_boundary(corrected_text, page_logs)
                        application_region_results.append({
                            'page_num': page_num,
                            'full_image': full_image,
                            'page_logs': page_logs,
                            'text': corrected_text,
                            'duration': duration,
                            'boundary': boundary,
                            'boundary_evidence': boundary_evidence,
                        })

                pf_thread_app.join(timeout=30)

                detected_boundaries = [item['page_num'] for item in application_region_results if item['boundary']]
                expected_boundaries = max(1, total_pages // PAGES_PER_CASE['申请书'])

                boundary_page_set = set(detected_boundaries)
                expected_start_pages = {
                    1 + index * PAGES_PER_CASE['申请书'] for index in range(expected_boundaries)
                }

                for item in application_region_results:
                    if cancel_check and cancel_check():
                        _log(f"  [申请书] 后处理阶段已取消")
                        raise CancelledError("用户取消")
                    page_num = item['page_num']
                    text = item['text']
                    page_logs = item['page_logs']
                    duration = item['duration']
                    method = 'region_first'
                    doc_ocr_app = _cfg.get_doc_ocr(doc_type)
                    needs_fallback, fallback_trigger_reason = _should_fallback_application(
                        page_num,
                        text,
                        detected_boundaries,
                        expected_boundaries,
                        expected_start_pages,
                        min_text_length=doc_ocr_app.region_fallback_min_text_length if doc_ocr_app else None,
                        page_logs=page_logs,
                    )

                    if needs_fallback and doc_ocr_app and doc_ocr_app.allow_full_page_fallback:
                        fallback_used = True
                        full_result = ocr.recognize_full_page_image(
                            item['full_image'],
                            page_num=page_num,
                            method='full_page_fallback',
                            optimize_output=True,
                        )
                        fallback_text = apply_ocr_corrections(full_result.text, doc_type=doc_type)
                        duration += full_result.duration
                        if fallback_text:
                            text = fallback_text
                            method = full_result.method

                    pages.append({
                        'page': page_num,
                        'text': text,
                        'method': method,
                        'duration': duration,
                        'region_attempts': page_logs,
                        'region_stats': _build_region_stats(page_logs),
                        'region_text_length': len(_compact_text(item['text'])),
                        'fallback_used': method == 'full_page_fallback',
                        'fallback_trigger_reason': fallback_trigger_reason if method == 'full_page_fallback' else None,
                        'boundary_detected': item['boundary'] or _is_application_boundary(text)[0],
                    })
            else:
                doc_cfg = _cfg.doc_type_map[doc_type]
                default_regions = _cfg.ocr_doc_regions.get(doc_type, [])

                all_page_nums = list(range(1, total_pages + 1))
                # 大文件使用批量预取，减少子进程开销
                # 377页PDF：逐页调用 = 377 × ~1.5s = 565s；批量20页 = 19批 × ~20s = 380s
                batch_size = 20 if total_pages > 50 else 5
                pf_queue: queue.Queue = queue.Queue(maxsize=max(4, batch_size // 2))
                pf_thread = threading.Thread(
                    target=_prefetch_pages_batch,
                    args=(region_extractor, pdf_path, all_page_nums, pf_queue, batch_size),
                    kwargs={'cancel_check': cancel_check},
                    daemon=True,
                )
                pf_thread.start()

                processed_count = 0
                while True:
                    prefetched = _get_prefetched(pf_queue, timeout=180)  # 3分钟超时
                    if prefetched is None:
                        break
                    page_num, full_image = prefetched
                    processed_count += 1
                    if isinstance(full_image, Exception):
                        _log(f"    [{doc_type}] 第{page_num}页提取失败: {full_image}")
                        continue
                    if cancel_check and cancel_check():
                        _log(f"    [{doc_type}] 已取消")
                        raise CancelledError("用户取消")
                    if page_num % 50 == 1 or page_num == total_pages:
                        _log(f"    [{doc_type}] {page_num}/{total_pages}...")
                    if page_progress and (processed_count % 20 == 0 or page_num == total_pages):
                        page_progress(page_num, total_pages)
                    primary_text, primary_logs = _collect_region_texts(
                        ocr,
                        region_extractor,
                        pdf_path,
                        page_num,
                        doc_type,
                        full_image=full_image,
                        region_names=default_regions,
                    )
                    page_logs = list(primary_logs)
                    combined_text = apply_ocr_corrections(primary_text, doc_type=doc_type)

                    if doc_type:
                        region_usable = _check_region_usable(doc_type, page_logs, combined_text)
                    else:
                        region_usable = len(combined_text.replace('\n', '').replace(' ', '')) >= 6

                    text = combined_text
                    method = 'region_first'
                    duration = sum(item['duration'] for item in page_logs)
                    marker_detected = bool(doc_cfg.content_marker and doc_cfg.content_marker in combined_text)

                    pages.append({
                        'page': page_num,
                        'text': text,
                        'method': method,
                        'duration': duration,
                        'region_attempts': page_logs,
                        'region_stats': _build_region_stats(page_logs),
                        'region_text_length': len(_compact_text(combined_text)),
                        'fallback_used': False,
                        'fallback_trigger_reason': None,
                        'marker_detected': marker_detected,
                        'region_usable': region_usable,
                    })

                pf_thread.join(timeout=30)

            total_duration = time.perf_counter() - total_start
            result = {
                'filename': pdf_path.name,
                'filepath': str(pdf_path),
                'total_pages': len(pages),
                'method': 'region_first' if not fallback_used else 'region_first_with_fallback',
                'total_duration': total_duration,
                'pages': pages,
                'full_text': "\n\n".join([f"=== 第{p['page']}页 ===\n{p['text']}" for p in pages if p['text']]),
                'performance_summary': _build_ocr_output_summary(pdf_path, pages, total_duration, doc_type=doc_type),
            }
        else:
            result = ocr.process_pdf(str(pdf_path), force_ocr=False)

        if result is None:
            _log(f"  [ERROR] OCR 识别失败: {pdf_path.name}")
            return {'pages': [], 'total_pages': 0, 'filename': pdf_path.name, 'method': 'error'}

        post_processor = post_processor or TextPostProcessor()
        processed_pages = []

        for page_data in result['pages']:
            text = page_data.get('text', '')
            text = apply_ocr_corrections(text, doc_type=doc_type)
            processed = post_processor.process(text)

            processed_pages.append({
                'page': page_data['page'],
                'text': processed['processed'],
                'original_text': text,
                'method': page_data.get('method', 'unknown'),
                'duration': page_data.get('duration', 0),
                'region_attempts': page_data.get('region_attempts', []),
                'region_stats': page_data.get('region_stats', _build_region_stats(page_data.get('region_attempts', []))),
                'region_text_length': page_data.get('region_text_length', 0),
                'fallback_used': page_data.get('fallback_used', False),
                'fallback_trigger_reason': page_data.get('fallback_trigger_reason'),
                'notice_matched': page_data.get('notice_matched', False),
                'boundary_detected': page_data.get('boundary_detected', False),
                'marker_detected': page_data.get('marker_detected', False),
                'region_usable': page_data.get('region_usable', False),
            })

        output = {
            'pages': processed_pages,
            'total_pages': result['total_pages'],
            'filename': result['filename'],
            'method': result['method'],
            'total_duration': result['total_duration'],
            'fallback_pages': sum(1 for page in processed_pages if page.get('fallback_used')),
            'region_pages': sum(1 for page in processed_pages if page.get('method') == 'region_first'),
            'optimization_strategy': result.get('method', 'unknown'),
            'stopped_early': result.get('stopped_early', False),
            'selected_notice': result.get('selected_notice'),
            'selected_page': result.get('selected_page'),
            'selected_root_notice': result.get('selected_root_notice'),
            'candidate_notices': result.get('candidate_notices', []),
            'performance_summary': result.get('performance_summary') or _build_ocr_output_summary(pdf_path, processed_pages, result.get('total_duration', 0), doc_type=doc_type),
        }

        perf = output.get('performance_summary', {})
        slowest_page = perf.get('slowest_page') or {}
        _log(
            f"  [OCR] 完成: {pdf_path.name} ({result['total_duration']:.2f}s, "
            f"fallback {output['fallback_pages']}/{output['total_pages']}, "
            f"最慢页 P{slowest_page.get('page', '-')} {slowest_page.get('duration', 0):.2f}s)"
        )
        return output

    except CancelledError:
        raise
    except Exception as e:
        _log(f"  [ERROR] OCR 处理异常: {pdf_path.name} - {e}")
        return {'pages': [], 'total_pages': 0, 'filename': pdf_path.name, 'method': 'error', 'error': str(e)}


def _run_ocr_with_timeout(pdf_path: Path, **kwargs) -> Dict:
    timeout = kwargs.pop('timeout', None)
    page_progress = kwargs.pop('page_progress', None)
    cancel_check = kwargs.get('cancel_check')
    kwargs['page_progress'] = page_progress
    if timeout is None:
        total_pages = inspect_pdf_page_count(pdf_path) if pdf_path.exists() else 1
        timeout = max(_PDF_TIMEOUT_SECONDS, total_pages * 5)
    poll_seconds = 0.25
    with ThreadPoolExecutor(max_workers=1) as executor:
        future = executor.submit(run_real_ocr_on_pdf, pdf_path, **kwargs)
        try:
            elapsed = 0.0
            while elapsed < timeout:
                try:
                    return future.result(timeout=poll_seconds)
                except FuturesTimeoutError:
                    elapsed += poll_seconds
                    if cancel_check and cancel_check():
                        _log(f"  [CANCEL] {pdf_path.name} 已取消")
                        raise CancelledError("用户取消")
            _log(f"  [TIMEOUT] {pdf_path.name} 处理超时 ({timeout}s)，跳过")
            return {'pages': [], 'total_pages': 0, 'filename': pdf_path.name, 'method': 'timeout'}
        except Exception as e:
            _log(f"  [ERROR] {pdf_path.name}: {e}")
            return {'pages': [], 'total_pages': 0, 'filename': pdf_path.name, 'method': 'error'}


# ---------------------------------------------------------------------------
# 流式批次处理器集成
# ---------------------------------------------------------------------------

_STREAMING_THRESHOLD = 50


def _should_use_streaming(total_tasks: int, actual_task_count: int = 0) -> bool:
    """流式路径缺少图片预处理、全页回退、责令号候选选择、空白页跳过等关键逻辑，
    识别准确率远低于串行路径。暂时禁用，后续补齐功能后再启用。"""
    return False


def _build_tasks_from_cases(input_dir: Path, cases: List[Dict]) -> List["Task"]:
    """从台账 cases 生成 Task 列表，供 StreamingBatchProcessor 消费。"""
    tasks: List["Task"] = []

    # 源 PDF 路径映射
    app_pdf = input_dir / _cfg.source_mapping.get('输出文件（申请书）', '申请书.pdf')
    auth_pdf = input_dir / _cfg.source_mapping.get('输出文件（授权书）', '授权书.pdf')
    letter_pdf = input_dir / _cfg.source_mapping.get('输出文件（所函）', '所函.pdf')

    # 授权书：作为一个整体 task（每页对应一个 case，OCR 后按页拆）
    if auth_pdf.exists():
        tasks.append(Task(
            task_id="auth_all",
            task_type='auth_all',
            source_file=str(auth_pdf),
            page_start=1,
            page_end=0,
            company_name=None,
            notice_number=None,
            sequence=None,
        ))

    # 所函：作为一个整体 task
    if letter_pdf.exists():
        tasks.append(Task(
            task_id="letter_all",
            task_type='letter_all',
            source_file=str(letter_pdf),
            page_start=1,
            page_end=0,
            company_name=None,
            notice_number=None,
            sequence=None,
        ))

    # 申请书：作为一个整体 task
    if app_pdf.exists():
        tasks.append(Task(
            task_id="application",
            task_type='application',
            source_file=str(app_pdf),
            page_start=1,
            page_end=0,
            company_name=None,
            notice_number=None,
            sequence=None,
        ))

    # 责催：每个存在的 notice PDF 对应一个 task
    notice_path_map = {path.name: path for path in iter_notice_pdf_paths(input_dir)}
    for name, path in notice_path_map.items():
        tasks.append(Task(
            task_id=f"notice_{name}",
            task_type='notice',
            source_file=str(path),
            page_start=1,
            page_end=1,
            company_name=None,
            notice_number=None,
            sequence=None,
        ))

    return tasks


def _streaming_db_fingerprint(input_dir: Path) -> str:
    """确定性指纹：基于输入目录路径 + 已有 PDF 文件名列表。"""
    import hashlib
    pdf_files = sorted(p.name for p in input_dir.glob('*.pdf') if p.is_file())
    raw = str(input_dir) + '\n' + '\n'.join(pdf_files)
    return hashlib.md5(raw.encode('utf-8')).hexdigest()[:12]


def _run_streaming_ocr(input_dir: Path, cases: List[Dict],
                       progress_callback: Optional[callable],
                       cancel_check: Optional[callable],
                       force: bool = False,
                       log_callback: Optional[callable] = None) -> Dict[str, Dict]:
    """流式 OCR 入口：构建任务 → 流式批次处理 → 聚合结果。"""
    if not HAS_STREAMING or TaskStateManager is None or StreamingBatchProcessor is None:
        _log("[WARN] 流式处理器未导入，回退到串行逻辑")
        return {}

    tasks = _build_tasks_from_cases(input_dir, cases)
    if not tasks:
        _log("[WARN] 无可用任务，流式处理器返回空结果")
        return {}

    fingerprint = _streaming_db_fingerprint(input_dir)
    db_path = USER_DATA_DIR / 'temp' / f'streaming_{fingerprint}.db'
    state_mgr = TaskStateManager(db_path)

    saved_fp = state_mgr.get_meta('input_fingerprint')
    if saved_fp and saved_fp != fingerprint:
        _log("输入文件已变化，清除旧缓存")
        state_mgr.clear_all()

    existing = state_mgr.get_summary()

    if existing['progress'] >= 1.0 and saved_fp == fingerprint and not force:
        _log(f"OCR 已完成（{existing['done']}/{existing['total']}），直接返回缓存结果")
        results = state_mgr.get_all_results()
        state_mgr.close()
        return results

    if force or existing['total'] == 0:
        state_mgr.clear_all()
        state_mgr.insert_tasks(tasks)
    else:
        _log(f"断点续跑: 已完成 {existing['done']}/{existing['total']}, 继续处理...")
        state_mgr.insert_tasks(tasks)

    state_mgr.set_meta('input_fingerprint', fingerprint)

    from non_litigation.streaming import set_log_fn as _set_streaming_log
    def _streaming_log(msg: str):
        if log_callback:
            try:
                log_callback("info", msg)
            except Exception:
                pass
        _log(msg)
    _set_streaming_log(_streaming_log)

    processor = StreamingBatchProcessor(state_mgr, batch_size=50, max_workers=1)
    processor.initialize()

    # 包装进度回调（summary dict -> 原有签名）
    def _streaming_progress(summary: Dict):
        if progress_callback:
            done = summary.get('done', 0)
            total = summary.get('total', 0)
            try:
                progress_callback(done, total, 'streaming_batch')
            except Exception:
                pass

    processor.run(progress_callback=_streaming_progress, cancel_check=cancel_check)

    # 聚合为 {filename: ocr_result} 格式，兼容现有调用方
    ocr_results = state_mgr.get_all_results()

    # 清理
    state_mgr.close()

    return ocr_results


def run_real_ocr(input_dir: Path, use_mock: bool = False,
                 progress_callback: Optional[callable] = None,
                 cancel_check: Optional[callable] = None,
                 log_callback: Optional[callable] = None,
                 cached_results: Optional[Dict[str, Dict]] = None,
                 force: bool = False,
                 result_callback: Optional[callable] = None,
                 task_output_dir: Optional[Path] = None) -> Dict[str, Dict]:
    ocr_start = time.perf_counter()
    ocr_results: Dict[str, Dict] = dict(cached_results) if cached_results else {}
    skipped_count = 0
    # P0-#extra: 提前到函数顶部初始化，让责催完成时也能写入
    type_durations: Dict[str, float] = {}
    notice_dur_captured: float = 0.0

    def _on_result(filename: str, result: Dict):
        if result_callback:
            try:
                result_callback(filename, result, ocr_results)
            except Exception:
                pass

    _state_dirty = False

    def _save_state():
        """标记状态为脏，不立即落盘。统一在 OCR 阶段结束时调用 _flush_state() 落盘。"""
        nonlocal _state_dirty
        _state_dirty = True

    def _flush_state():
        """把内存中的 ocr_results 持久化到磁盘"""
        nonlocal _state_dirty
        if not _state_dirty or task_output_dir is None:
            return
        try:
            debug_dir = task_output_dir / "debug"
            debug_dir.mkdir(parents=True, exist_ok=True)
            state_path = debug_dir / "ocr_state.json"
            state = {
                'completed': list(ocr_results.keys()),
                'timestamp': time.time(),
            }
            state_path.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding='utf-8')
            _state_dirty = False
        except Exception as e:
            _log(f"  [WARN] 保存断点状态失败: {e}")

    def _load_state() -> set:
        """从磁盘加载已完成任务状态"""
        try:
            state_path = task_output_dir / "debug" / "ocr_state.json"
            # 兼容旧版路径
            if not state_path.exists():
                state_path = task_output_dir / ".ocr_state.json"
            if not state_path.exists():
                return set()
            data = json.loads(state_path.read_text(encoding='utf-8'))
            return set(data.get('completed', []))
        except Exception as e:
            _log(f"  [WARN] 加载断点状态失败: {e}")
            return set()

    def _compute_eta(done_count: int, total_count: int) -> str:
        """计算预估剩余时间（ETA）"""
        if done_count == 0 or total_count == 0:
            return ""
        elapsed = time.perf_counter() - ocr_start
        avg_per_task = elapsed / done_count
        remaining = (total_count - done_count) * avg_per_task
        if remaining < 60:
            return f"ETA {remaining:.0f}s"
        elif remaining < 3600:
            return f"ETA {int(remaining // 60)}m{int(remaining % 60)}s"
        else:
            return f"ETA {int(remaining // 3600)}h{int((remaining % 3600) // 60)}m"

    def _report(msg: str):
        _log(msg)
        if log_callback:
            log_callback("info", msg)

    _report(f"OCR 引擎状态: HAS_OCR={HAS_OCR}, use_mock={use_mock}")

    # ---------- 断点续跑：加载已完成任务 ----------
    if task_output_dir is not None and not force:
        persisted = _load_state()
        if persisted:
            _report(f"[断点续跑] 发现已完成任务 {len(persisted)} 个，自动跳过")
            for fn in persisted:
                if fn not in ocr_results:
                    ocr_results[fn] = {'pages': [], 'total_pages': 0, 'filename': fn, 'method': 'resumed'}

    # ---------- 自动流式切换判断 ----------
    total_tasks = 0
    cases: List[Dict] = []
    try:
        # 尝试从 input_dir 的父目录加载台账（常规目录结构）
        sample_root = input_dir.parent
        cases = load_non_litigation_cases(sample_root)
        # 每个 case 约产生 2 个细粒度任务(auth/letter) + 1 个 application + notice
        total_tasks = len(cases) * 2 + 1 + len(discover_notice_files(input_dir))
    except Exception:
        # 加载失败时回退到文件数统计
        notice_files = discover_notice_files(input_dir)
        other_files_expected = [
            (pdf_name, pdf_name.replace('.pdf', ''))
            for pdf_name in _cfg.source_mapping.values()
        ]
        other_existing = [(name, stem) for name, stem in other_files_expected if (input_dir / name).exists()]
        total_tasks = len(notice_files) + len(other_existing)
        cases = []

    if cases and _should_use_streaming(total_tasks, len(_build_tasks_from_cases(input_dir, cases))) and not use_mock and HAS_OCR and HAS_STREAMING:
        _report(f"总任务数 {total_tasks} >= {_STREAMING_THRESHOLD}，启用流式批次处理器")
        streaming_results = _run_streaming_ocr(input_dir, cases, progress_callback, cancel_check, force, log_callback)
        if cached_results:
            merged = dict(cached_results)
            merged.update(streaming_results)
            ocr_results = merged
        else:
            ocr_results = streaming_results
        total_duration = round(time.perf_counter() - ocr_start, 4)
        _report(f"OCR 阶段完成(流式): {total_duration:.2f}s, 共处理 {len(ocr_results)} 个文件")
        _flush_state()  # 流式模式结束，一次性落盘
        return ocr_results
    # -------------------------------------

    notice_files = discover_notice_files(input_dir)
    notice_path_map = {path.name: path for path in iter_notice_pdf_paths(input_dir)}

    other_files_expected = [
        (pdf_name, pdf_name.replace('.pdf', ''))
        for pdf_name in _cfg.source_mapping.values()
    ]
    other_existing = [(name, stem) for name, stem in other_files_expected if (input_dir / name).exists()]

    total_found = len(notice_files) + len(other_existing)
    _report(f"文件发现: 责催 {len(notice_files)} 个, 其他 {len(other_existing)} 个, 合计 {total_found} 个")

    _global_done = [0]
    _global_total = total_found

    def _progress_update(filename: str):
        _global_done[0] += 1
        if progress_callback:
            progress_callback(_global_done[0], _global_total, filename)

    shared_ocr = None
    shared_region_extractor = None
    shared_post_processor = None

    notice_items = []
    for source_name in notice_files:
        pdf_path = notice_path_map.get(source_name, input_dir / source_name)
        if pdf_path.exists():
            notice_items.append((source_name, pdf_path))
        else:
            _report(f"  [WARN] 文件不存在: {pdf_path}")

    if notice_items and not use_mock and HAS_OCR:
        notice_count = len(notice_items)
        notice_pending = [(n, p) for n, p in notice_items if n not in ocr_results]
        skipped = notice_count - len(notice_pending)
        if skipped:
            _report(f"处理责催文件（{len(notice_pending)} 个新文件，跳过 {skipped} 个缓存）...")
        else:
            _report(f"处理责催文件（逐页识别，找到即停）... 共 {notice_count} 个")

        shared_ocr, shared_region_extractor = _build_ocr_processors()
        shared_post_processor = TextPostProcessor()

        from core.pdf_ocr_ultra import detect_gpu_provider
        gpu_provider, _ = detect_gpu_provider()
        can_parallel = gpu_provider not in ('dml_det',) and len(notice_pending) > 1
        notice_workers = min(len(notice_pending), 3) if can_parallel else 1
        notice_start = time.perf_counter()
        if can_parallel:
            _report(f"  [并行] 责催启用 {notice_workers} 线程 (GPU={gpu_provider})")

        def _ocr_notice_one(task):
            source_name, pdf_path = task
            t0 = time.perf_counter()
            try:
                result = _run_ocr_with_timeout(
                    pdf_path,
                    use_mock=use_mock,
                    is_notice=True,
                    stop_pattern=NOTICE_PATTERN,
                    doc_type='责催',
                    ocr=shared_ocr,
                    region_extractor=shared_region_extractor,
                    post_processor=shared_post_processor,
                    cancel_check=cancel_check,
                )
            except CancelledError:
                return source_name, None, 'cancelled'
            except Exception as e:
                _report(f"  [ERROR] 责催 {source_name}: {e}")
                return source_name, {'pages': [], 'total_pages': 0, 'filename': source_name, 'method': 'error'}, time.perf_counter() - t0
            return source_name, result, time.perf_counter() - t0

        if notice_workers > 1:
            with ThreadPoolExecutor(max_workers=notice_workers) as pool:
                future_map = {pool.submit(_ocr_notice_one, t): t[0] for t in notice_pending}
                done_idx = 0
                for future in as_completed(future_map):
                    if cancel_check and cancel_check():
                        for f in future_map:
                            f.cancel()
                        _report(f"责催已取消，已完成 {done_idx}/{len(notice_pending)}")
                        raise CancelledError("用户取消")
                    source_name, result, dur = future.result()
                    if result is None:
                        continue  # cancelled
                    done_idx += 1
                    ocr_results[source_name] = result
                    _on_result(source_name, result)
                    _save_state()
                    _report(f"[{done_idx}/{len(notice_pending)}] {source_name} ({dur:.1f}s)  {_compute_eta(done_idx + skipped, notice_count)}")
                    _progress_update(source_name)
        else:
            for idx, (source_name, pdf_path) in enumerate(notice_pending, 1):
                if cancel_check and cancel_check():
                    _report(f"责催已取消，已完成 {idx-1}/{len(notice_pending)}")
                    raise CancelledError("用户取消")
                source_name, result, file_dur = _ocr_notice_one((source_name, pdf_path))
                if result is None:
                    continue
                ocr_results[source_name] = result
                _on_result(source_name, result)
                _save_state()
                _report(f"[{idx}/{len(notice_pending)}] {source_name} ({file_dur:.1f}s)  {_compute_eta(idx + skipped, notice_count)}")
                _progress_update(source_name)

        notice_dur = time.perf_counter() - notice_start
        _report(f"责催完成: {len(notice_pending)} 个文件, 耗时 {notice_dur:.1f}s")
        type_durations['责催'] = round(notice_dur, 4)
        notice_dur_captured = round(notice_dur, 4)
    elif notice_items and use_mock:
        for source_name, pdf_path in notice_items:
            if source_name not in ocr_results:
                ocr_results[source_name] = _mock_ocr_result(source_name)
                _on_result(source_name, ocr_results[source_name])
                _progress_update(source_name)

    if cancel_check and cancel_check():
        _report(f"任务已取消，跳过其他文件处理，已缓存 {len(ocr_results)} 个结果")
        raise CancelledError("用户取消")

    other_files = [
        (pdf_name, pdf_name.replace('.pdf', ''))
        for pdf_name in _cfg.source_mapping.values()
    ]

    _report("处理其他文件...")
    if not use_mock and HAS_OCR and shared_ocr is None:
        shared_ocr, shared_region_extractor = _build_ocr_processors()
        shared_post_processor = TextPostProcessor()

    parallel_candidates = []
    serial_candidates = []
    for filename, stem in other_files:
        pdf_path = input_dir / filename
        if not pdf_path.exists():
            continue
        if filename in ocr_results:
            skipped_count += 1
            _progress_update(filename)
            _report(f"[SKIP] {filename} (已有缓存)")
            continue
        doc_type = stem if stem in {'申请书', '授权书', '所函'} else None
        if doc_type:
            parallel_candidates.append((filename, doc_type, pdf_path))
        else:
            serial_candidates.append((filename, stem, pdf_path))

    if parallel_candidates:
        from core.pdf_ocr_ultra import detect_gpu_provider
        gpu_provider, _ = detect_gpu_provider()
        use_parallel = gpu_provider not in ('dml_det',) and len(parallel_candidates) > 1

        t_parallel = time.perf_counter()

        if use_parallel:
            _report(f"  [并行] 启动 {len(parallel_candidates)} 个文档的多线程 OCR (GPU={gpu_provider})...")

            def _ocr_one_file(task):
                filename, doc_type, pdf_path = task
                t0 = time.perf_counter()
                try:
                    res = _run_ocr_with_timeout(
                        pdf_path,
                        use_mock=use_mock,
                        doc_type=doc_type,
                        ocr=shared_ocr,
                        region_extractor=shared_region_extractor,
                        post_processor=shared_post_processor,
                        cancel_check=cancel_check,
                    )
                except CancelledError:
                    return filename, None, 0.0
                except Exception as e:
                    _report(f"  [并行失败] {filename}: {e}")
                    return filename, {'pages': [], 'total_pages': 0, 'filename': filename, 'method': 'error'}, time.perf_counter() - t0
                return filename, res, time.perf_counter() - t0

            from concurrent.futures import as_completed
            with ThreadPoolExecutor(max_workers=min(len(parallel_candidates), 3)) as pool:
                future_map = {
                    pool.submit(_ocr_one_file, task): task[0]
                    for task in parallel_candidates
                }
                done_idx = 0
                for future in as_completed(future_map):
                    filename = future_map[future]
                    try:
                        _, result, dur = future.result()
                    except CancelledError:
                        raise
                    except Exception as e:
                        _report(f"  [并行失败] {filename}: {e}")
                        continue
                    if result is None:
                        continue
                    done_idx += 1
                    ocr_results[filename] = result
                    _on_result(filename, result)
                    _save_state()
                    eta = _compute_eta(done_idx + skipped_count, len(parallel_candidates))
                    _report(f"  [并行 {done_idx}/{len(parallel_candidates)}] {filename} ({dur:.1f}s)  {eta}")
                    _progress_update(filename)
        else:
            _report(f"  [串行] DirectML 模式下禁止多线程并行，改为逐文件处理...")
            for idx, (filename, doc_type, pdf_path) in enumerate(parallel_candidates, 1):
                if cancel_check and cancel_check():
                    _report(f"  [取消] 已取消，已完成 {idx-1}/{len(parallel_candidates)}")
                    raise CancelledError("用户取消")
                total_pages = inspect_pdf_page_count(pdf_path) if pdf_path.exists() else 0
                _report(f"  [{idx}/{len(parallel_candidates)}] {filename} ({total_pages} pages)...")
                t_file = time.perf_counter()
                result = _run_ocr_with_timeout(
                    pdf_path,
                    use_mock=use_mock,
                    doc_type=doc_type,
                    ocr=shared_ocr,
                    region_extractor=shared_region_extractor,
                    post_processor=shared_post_processor,
                    cancel_check=cancel_check,
                )
                ocr_results[filename] = result
                _on_result(filename, result)
                _save_state()
                file_dur = time.perf_counter() - t_file
                eta = _compute_eta(idx + skipped_count, len(parallel_candidates))
                _report(f"  [{idx}/{len(parallel_candidates)}] {filename} done ({file_dur:.1f}s)  {eta}")
                _progress_update(filename)
                # P0-#extra: dml_det 串行分支累加 type_durations
                type_durations[doc_type or '其他'] = type_durations.get(doc_type or '其他', 0.0) + file_dur

        parallel_dur = time.perf_counter() - t_parallel
        mode_label = '并行' if use_parallel else '串行'
        _report(f"  [{mode_label}] 完成: {len(parallel_candidates)} 个文件, 耗时 {parallel_dur:.1f}s")

    other_idx = len(parallel_candidates)
    other_total = len(parallel_candidates) + len(serial_candidates)
    # type_durations 已在 run_real_ocr 顶部初始化（line 1905），不再覆盖
    for filename, stem, pdf_path in serial_candidates:
        if cancel_check and cancel_check():
            _report(f"其他文件已取消，已完成 {other_idx}/{other_total}")
            raise CancelledError("用户取消")
        other_idx += 1
        total_pages = inspect_pdf_page_count(pdf_path)
        _report(f"[{other_idx}/{other_total}] {filename} ({total_pages} pages)...")
        t_file = time.perf_counter()

        def _make_page_progress(fname: str, tot: int):
            def _cb(done: int, total: int):
                _report(f"  {fname}: {done}/{total} pages")
            return _cb

        result = _run_ocr_with_timeout(
            pdf_path,
            use_mock=use_mock,
            doc_type=None,
            ocr=shared_ocr,
            region_extractor=shared_region_extractor,
            post_processor=shared_post_processor,
            page_progress=_make_page_progress(filename, total_pages),
            cancel_check=cancel_check,
        )
        file_dur = time.perf_counter() - t_file
        ocr_results[filename] = result
        _on_result(filename, result)
        _save_state()
        eta = _compute_eta(other_idx, other_total)
        _report(f"[{other_idx}/{other_total}] {filename} 完成 ({file_dur:.1f}s)  {eta}")
        _progress_update(filename)

    for dt, dur in type_durations.items():
        _report(f"{dt}完成: {dur:.1f}s")

    total_duration = round(time.perf_counter() - ocr_start, 4)
    if skipped_count:
        _report(f"跳过缓存: {skipped_count} 个文件")
    _report(f"OCR 阶段完成: {total_duration:.2f}s, 共处理 {len(ocr_results)} 个文件")
    _flush_state()  # OCR 阶段结束，一次性落盘断点状态

    # P0-#extra: 暴露 OCR 阶段计时明细，让 build_run_summary 能聚合到 phase_timings。
    if task_output_dir is not None:
        try:
            timing = {
                'total_seconds': total_duration,
                'skipped_count': skipped_count,
                'type_durations': dict(type_durations),
                'file_count': len(ocr_results),
            }
            # _meta 不会进入导出环节的 .pdf 拷贝
            ocr_results['_meta'] = {'ocr_timing': timing}
            meta_path = task_output_dir / 'debug' / 'ocr_timing.json'
            meta_path.parent.mkdir(parents=True, exist_ok=True)
            meta_path.write_text(
                __import__('json').dumps(timing, ensure_ascii=False, indent=2),
                encoding='utf-8',
            )
        except Exception as e:
            _log(f"  [WARN] 保存 OCR 阶段计时失败: {e}")

    return ocr_results


def export_pdf_ranges(source_pdf: Path, ranges: List[Tuple[int, int]], output_dir: Path, target_names: List[str]) -> int:
    created = 0
    with open_pdf_reader(source_pdf) as reader:
        for i, ((start, end), target_name) in enumerate(zip(ranges, target_names)):
            target_path = output_dir / target_name
            if target_path.exists():
                created += 1
                _log(f"  [SKIP] 跳过已存在: {target_name}")
                continue

            if start >= len(reader.pages):
                _log(f"  [ERROR] 页码超出范围: {target_name} (起始页 {start} >= 总页数 {len(reader.pages)})")
                continue

            writer = PdfWriter()
            actual_end = min(end, len(reader.pages))
            for page_index in range(start, actual_end):
                writer.add_page(reader.pages[page_index])

            with target_path.open('wb') as file_obj:
                writer.write(file_obj)

            created += 1
            _log(f"  [OK] 导出: {target_name} (第 {start+1}-{actual_end} 页)")

    return created


def export_notice_files(sample_root: Path, input_dir: Path, output_dir: Path, ocr_results: Dict[str, Dict], excel_path: Optional[Path] = None, error_root: Optional[Path] = None) -> int:
    cases = load_non_litigation_cases(sample_root, excel_path=excel_path)

    target_map = {}
    target_notice_map = {}
    ledger_notices = []
    for case in cases:
        normalized = normalize_notice_number(case['notice_number'])
        target_name = f"{case['sequence']}-责催-{case['notice_number']}.pdf"
        target_map[normalized] = target_name
        target_notice_map[target_name] = normalized
        ledger_notices.append(case['notice_number'])

    # 延迟创建错误文件夹（只在有需要核查的文件时才创建）
    error_dir: Optional[Path] = None
    def _get_error_dir() -> Path:
        nonlocal error_dir, error_root
        if error_dir is None:
            if error_root is None:
                error_root = output_dir.parent / '需人工核查'
            error_dir = error_root / '责催'
            error_dir.mkdir(parents=True, exist_ok=True)
        return error_dir

    notice_files = discover_notice_files(input_dir)
    # 传入台账数据进行 OCR 识别结果比对纠错（包含智能纠错）
    source_map = detect_notice_source_mapping_from_ocr(
        ocr_results, notice_files,
        ledger_notices=ledger_notices,
        notice_to_target=target_map,
        ledger_cases=cases,  # 传入完整台账数据用于智能纠错
    )

    created = 0
    unmatched = []

    for source_name, detected_notice in source_map.items():
        # P0-#9: detected_notice 来自 _match_notice_with_ledger，已经是 ledger 字符串；
        # 拿"OCR 实际识别"得从 ocr_results[source_name].selected_notice 取原始 OCR 输出。
        ocr_selected_notice = (ocr_results.get(source_name) or {}).get('selected_notice') or detected_notice
        target_name = target_map.get(detected_notice)
        detected_root = _get_notice_root_number(detected_notice)
        ratio = None

        if not target_name and detected_root in target_map:
            target_name = target_map[detected_root]
            match_type = 'same_root_base'
            _log_audit('root_base_match', {
                'source': source_name,
                'detected': detected_notice,
                'matched_target': target_name,
                'root_notice': detected_root,
            })
            _log(f"    [同根主号匹配] '{detected_notice}' -> '{target_name}'")
        elif target_name:
            match_type = 'exact'
        else:
            target_name, ratio = fuzzy_match_notice(detected_notice, target_map)
            if target_name:
                match_type = 'fuzzy'
                _log_audit('fuzzy_match', {
                    'source': source_name,
                    'detected': detected_notice,
                    'matched_target': target_name,
                    'similarity': ratio,
                })
                _log(f"    [INFO] 模糊匹配: '{detected_notice}' -> '{target_name}' (相似度: {ratio:.1%})")
                _log(f"    [WARN] 模糊匹配需人工确认！已记录到审计日志")

        if target_name:
            target_notice = target_notice_map.get(target_name)
            same_root_remap = bool(
                target_notice
                and target_notice != detected_notice
                and _get_notice_root_number(target_notice) == detected_root
            )
            export_metadata = {
                'matched_target': target_name,
                'matched_target_notice': target_notice,
                'export_match_type': match_type,
                'same_root_remap': same_root_remap,
            }
            if ratio is not None:
                export_metadata['export_match_similarity'] = ratio
            if same_root_remap:
                export_metadata['same_root_selected_notice'] = detected_notice
                _log_audit('same_root_remap', {
                    'source': source_name,
                    'selected_notice': detected_notice,
                    'target_notice': target_notice,
                    'matched_target': target_name,
                    'match_type': match_type,
                })
                _log(f"    [WARN] 主号识别后按同根目标导出: '{detected_notice}' -> '{target_notice}'")
            src = next((path for path in iter_notice_pdf_paths(input_dir) if path.name == source_name), input_dir / source_name)
            dst = output_dir / target_name
            # P0-#9: same_root_remap 时把 OCR 实际识别到的主号附加到文件名后缀；
            # dst 已存在则自动加 -1/-2 后缀，绝不覆盖既有产出。
            if same_root_remap and ocr_selected_notice and ocr_selected_notice != target_notice:
                dst_name = _append_detected_notice_suffix(target_name, ocr_selected_notice)
                dst = output_dir / dst_name
                _log(f"    [HINT] 同根主号匹配，文件名附加 OCR 实际识别号: {target_name} -> {dst_name}")
            if dst.exists():
                dst = _dedupe_destination(dst)
                _log(f"  [WARN] 目标已存在，自动加后缀: {dst.name}")
            if src.exists():
                shutil.copy2(src, dst)
                created += 1
                _log(f"  [OK] {source_name} -> {dst.name}")
                _log_audit('notice_renamed', {
                    'source': source_name,
                    'target': dst.name,
                    'match_type': match_type,
                    'detected_notice': detected_notice,
                    'target_notice': target_notice,
                    'same_root_remap': same_root_remap,
                })
            else:
                _log(f"  [ERROR] 源文件不存在: {src}")
                unmatched.append((source_name, detected_notice, "源文件不存在"))
        else:
            _log(f"  [ERROR] 无法匹配: {source_name} (识别到 '{detected_notice}')")
            unmatched.append((source_name, detected_notice, "无匹配台账"))
            _log_audit('match_failed', {
                'source': source_name,
                'detected': detected_notice,
            })
            
            # 复制到错误文件夹，添加识别信息到文件名
            src = next((path for path in iter_notice_pdf_paths(input_dir) if path.name == source_name), input_dir / source_name)
            if src.exists():
                # 延迟创建错误文件夹
                err_dir = _get_error_dir()
                # 构建错误文件名：原文件名_识别到[责令号]_需人工核查.pdf
                safe_detected = detected_notice.replace('/', '_').replace('\\', '_') if detected_notice else '未识别'
                error_filename = f"{Path(source_name).stem}_识别到[{safe_detected}]_需人工核查.pdf"
                error_dst = err_dir / error_filename
                
                # 如果文件已存在，添加序号
                counter = 1
                original_error_dst = error_dst
                while error_dst.exists():
                    stem = original_error_dst.stem
                    error_dst = err_dir / f"{stem}_{counter}.pdf"
                    counter += 1
                
                shutil.copy2(src, error_dst)
                _log(f"  [MOVED] 已复制到错误文件夹: {error_filename}")
                _log_audit('notice_moved_to_error', {
                    'source': source_name,
                    'detected': detected_notice,
                    'error_path': str(error_dst),
                })

    if unmatched:
        err_dir_path = _get_error_dir() if error_dir else (error_root / '责催' if error_root else output_dir.parent / '需人工核查' / '责催')
        _log(f"\n  [WARN] 未匹配文件汇总 ({len(unmatched)} 个)，已复制到: {err_dir_path}")
        for source_name, notice, reason in unmatched:
            _log(f"    - {source_name}: {reason} (识别: '{notice}')")

    return created


def export_application_files(input_dir: Path, output_dir: Path, target_names: List[str], ocr_results: Dict[str, Dict], error_root: Optional[Path] = None) -> int:
    source_pdf = input_dir / SOURCE_MAPPING['输出文件（申请书）'] 

    if not source_pdf.exists():
        _log(f"  [ERROR] 申请书文件不存在: {source_pdf}")
        return 0

    total_pages = inspect_pdf_page_count(source_pdf)
    expected_cases = len(target_names)

    _log(f"  [INFO] 申请书: {total_pages} 页，台账期望 {expected_cases} 个案件")

    # 延迟创建错误文件夹（只在有需要核查的文件时才创建）
    error_dir: Optional[Path] = None
    def _get_error_dir() -> Path:
        nonlocal error_dir, error_root
        if error_dir is None:
            if error_root is None:
                error_root = output_dir.parent / '需人工核查'
            error_dir = error_root / '申请书'
            error_dir.mkdir(parents=True, exist_ok=True)
        return error_dir

    ranges = detect_application_page_ranges_by_ocr(ocr_results, total_pages, expected_cases)

    if len(ranges) != expected_cases:
        _log(f"  [INFO] 按实际识别到 {len(ranges)} 个案件处理（台账 {expected_cases} 个）")

    # 构建台账责令号查找表
    post_processor = TextPostProcessor()
    notice_to_target: Dict[str, str] = {}
    ledger_notices: List[str] = []
    for target_name in target_names:
        stem = Path(target_name).stem
        parts = stem.split('-申请书pdf-', 1)
        if len(parts) == 2:
            notice_str = parts[1]
            normalized = normalize_notice_number(notice_str)
            if normalized not in notice_to_target:
                notice_to_target[normalized] = target_name
                ledger_notices.append(notice_str)

    # 提取 OCR 文本并为每个页面范围匹配责令号
    data = _get_ocr_result(ocr_results, '申请书')
    normal_ranges: List[Tuple[int, int]] = []
    normal_names: List[str] = []
    error_ranges: List[Tuple[int, int]] = []
    error_names: List[str] = []
    
    ledger_match_count = 0
    fuzzy_match_count = 0
    error_count = 0

    if data and data.get('pages'):
        for range_idx, (start_page, end_page) in enumerate(ranges):
            case_texts = []
            for page_data in data['pages']:
                page_num = page_data.get('page', 0)
                if start_page < page_num <= end_page:
                    text = page_data.get('text', '')
                    if text:
                        case_texts.append(text)

            combined = '\n'.join(case_texts)

            # 三级匹配：严格正则 → 宽松正则+模糊匹配 → 结构化纠错
            final_notice = _match_notice_from_ocr_text(
                combined, post_processor, notice_to_target, ledger_notices
            )
            src_label = final_notice.get('source', 'unknown')
            notice_number = final_notice.get('notice_number')
            target_name_matched = final_notice.get('target_name')
            needs_review = final_notice.get('needs_review', True)

            if target_name_matched and not needs_review:
                # 成功匹配台账（高置信度），正常导出
                normal_ranges.append((start_page, end_page))
                normal_names.append(target_name_matched)
                ledger_match_count += 1
                _log(f"  [OK] 申请书 第{start_page+1}-{end_page}页: '{notice_number}' -> '{target_name_matched}' ({src_label})")
            elif target_name_matched and needs_review:
                # 匹配成功但需要人工核查（模糊匹配等）
                error_ranges.append((start_page, end_page))
                safe_notice = notice_number.replace('/', '_').replace('\\', '_')
                error_name = f"申请书_第{start_page+1}-{end_page}页_识别到[{safe_notice}]_匹配方式[{src_label}]_需人工核查.pdf"
                error_names.append(error_name)
                fuzzy_match_count += 1
                error_count += 1
                _log(f"  [REVIEW] 申请书 第{start_page+1}-{end_page}页: '{notice_number}' -> '{target_name_matched}' ({src_label}) - 需人工核查")
                _log_audit('application_needs_review', {
                    'pages': f"{start_page+1}-{end_page}",
                    'detected_notice': notice_number,
                    'matched_target': target_name_matched,
                    'match_source': src_label,
                    'reason': '模糊匹配或结构化纠错，需要人工确认'
                })
            elif notice_number:
                # 识别到责令号但未匹配台账，导出到错误文件夹
                error_ranges.append((start_page, end_page))
                safe_notice = notice_number.replace('/', '_').replace('\\', '_')
                error_name = f"申请书_第{start_page+1}-{end_page}页_识别到[{safe_notice}]_未匹配台账_需人工核查.pdf"
                error_names.append(error_name)
                fuzzy_match_count += 1
                error_count += 1
                _log(f"  [WARN] 申请书 第{start_page+1}-{end_page}页: '{notice_number}' 未匹配台账，将导出到错误文件夹")
                _log_audit('application_unmatched', {
                    'pages': f"{start_page+1}-{end_page}",
                    'detected_notice': notice_number,
                    'reason': '未匹配台账'
                })
            else:
                # 未识别到责令号，导出到错误文件夹
                error_ranges.append((start_page, end_page))
                error_name = f"申请书_第{start_page+1}-{end_page}页_未识别到责令号_需人工核查.pdf"
                error_names.append(error_name)
                error_count += 1
                _log(f"  [WARN] 申请书 第{start_page+1}-{end_page}页: 未识别到责令号，将导出到错误文件夹")
                _log_audit('application_unknown', {
                    'pages': f"{start_page+1}-{end_page}",
                    'reason': '未识别到责令号'
                })

    # 导出正常匹配的文件
    created = 0
    if normal_ranges:
        _log(f"  [INFO] 申请书: 台账匹配 {ledger_match_count} 个，导出到正常文件夹")
        created += export_pdf_ranges(source_pdf, normal_ranges, output_dir, normal_names)
    
    # 导出匹配失败的文件到错误文件夹
    if error_ranges:
        _log(f"  [INFO] 申请书: 匹配失败 {error_count} 个，导出到错误文件夹")
        err_dir = _get_error_dir()
        created += export_pdf_ranges(source_pdf, error_ranges, err_dir, error_names)

    if not data or not data.get('pages'):
        _log(f"  [WARN] 无 OCR 数据，按台账顺序命名（可能不准确）")
        return export_pdf_ranges(source_pdf, ranges, output_dir, target_names)

    return created


def _extract_company_name_from_target(target_name: str) -> str:
    stem = Path(target_name).stem
    return normalize_company_name_for_matching(stem)


_COMPANY_SUFFIXES = '有限公司|股份有限公司|有限责任公司|集团公司|合伙企业|分公司|幼儿园|事务所|研究院|服务中心|合作社|医院|大学|学院|幼儿园'
_EXECUTION_PATTERN = re.compile(r'与(.+?)(?:关于|强制执行)')
_AGENCY_PATTERN = re.compile(r'与(.+?)关于.*?(?:执行|审查|一案)')

_COMPANY_STRIP_PREFIXES = re.compile(r'^(委托人|申请人|被执行人|申请执行人)[:：]?')
_COMPANY_STRIP_SUFFIXES = re.compile(r'(?:关于|强制执行|行政非诉审查|补缴).*$')

# 公司类型后缀，用于区分总公司/分公司/有限公司/股份有限公司等
_COMPANY_TYPE_SUFFIXES = [
    '股份有限公司', '有限责任公司', '有限公司', '集团公司', '股份公司',
    '合伙企业', '分公司', '事务所', '研究院', '服务中心', '合作社',
    '医院', '大学', '学院', '幼儿园', '管理中心', '协会', '中心',
]


def _extract_company_core_name(name: str) -> str:
    """提取公司核心名称（去掉通用后缀）"""
    name = name.strip()
    for suffix in _COMPANY_TYPE_SUFFIXES:
        if name.endswith(suffix):
            return name[:-len(suffix)].strip()
    return name


def _get_company_suffixes(name: str) -> List[str]:
    """提取公司名称中的所有类型后缀"""
    result = []
    name = name.strip()
    for suffix in _COMPANY_TYPE_SUFFIXES:
        if suffix in name:
            result.append(suffix)
    return result


def _extract_target_company(text: str, fallback_fn=None) -> Optional[str]:
    collapsed = re.sub(r'\n', '', text)
    m = _EXECUTION_PATTERN.search(collapsed)
    if not m:
        m = _AGENCY_PATTERN.search(collapsed)
    if m:
        name = m.group(1)
        name = _COMPANY_STRIP_PREFIXES.sub('', name).strip()
        if len(name) >= 4:
            return name
    if fallback_fn:
        return fallback_fn(text)
    return None


def _is_suffix_compatible(detected: str, target: str) -> bool:
    """
    检查两个公司名的后缀是否兼容，避免总公司匹配到分公司。
    核心规则：如果一方明确包含'分公司'，另一方不能是不含'分公司'的总公司。
    """
    detected_suffixes = set(_get_company_suffixes(detected))
    target_suffixes = set(_get_company_suffixes(target))

    # 如果一方有"分公司"另一方没有，不兼容
    has_branch_d = '分公司' in detected_suffixes
    has_branch_t = '分公司' in target_suffixes
    if has_branch_d != has_branch_t:
        return False

    return True


def _fuzzy_match_company_name(detected: str, target_names: List[str], used_indices: set,
                               threshold: float = 0.85) -> Optional[Tuple[int, str]]:
    """
    公司名称模糊匹配，策略：
    1. 精确匹配（规范化后完全一致）
    2. 核心名称匹配（去掉通用后缀后一致，且后缀兼容）
    3. SequenceMatcher 模糊匹配（≥阈值，且后缀兼容）
    """
    best_idx = None
    best_target = None
    best_score = -1

    detected_norm = normalize_company_name_for_matching(detected)
    detected_core = _extract_company_core_name(detected_norm)

    candidates = []
    for i, target in enumerate(target_names):
        if i in used_indices:
            continue
        target_company = _extract_company_name_from_target(target)
        target_norm = normalize_company_name_for_matching(target_company)

        # 1. 精确匹配
        if detected_norm == target_norm:
            candidates.append((i, target, 100.0, 'exact'))
            continue

        # 2. 核心名称匹配（去掉有限公司/分公司等后缀）
        target_core = _extract_company_core_name(target_norm)
        if detected_core and target_core and detected_core == target_core:
            if _is_suffix_compatible(detected_norm, target_norm):
                candidates.append((i, target, 95.0, 'core'))
                continue

        # 3. 模糊匹配
        ratio = SequenceMatcher(None, detected_norm, target_norm).ratio()
        if ratio >= threshold and _is_suffix_compatible(detected_norm, target_norm):
            candidates.append((i, target, ratio, 'fuzzy'))

    if not candidates:
        return None

    # 排序：精确 > 核心 > 模糊；同类型选更长的（更具体）
    def sort_key(item):
        score, match_type, name = item[2], item[3], item[1]
        type_order = {'exact': 3, 'core': 2, 'fuzzy': 1}
        return (type_order.get(match_type, 0), score, len(name))

    candidates.sort(key=sort_key, reverse=True)
    best = candidates[0]
    return (best[0], best[1])


def detect_company_page_mapping_from_ocr(ocr_results: Dict[str, Dict], doc_type: str,
                                          target_names: List[str]) -> Optional[List[str]]:
    data = _get_ocr_result(ocr_results, doc_type)
    if not data or not data.get('pages'):
        return None

    post_processor = TextPostProcessor()
    pages = data['pages']
    if len(pages) < len(target_names):
        return None

    matched_targets: List[Optional[str]] = [None] * len(pages)
    used_indices: set = set()

    for page_idx, page_data in enumerate(pages):
        text = page_data.get('text', '')
        detected_company = _extract_target_company(text, fallback_fn=post_processor.extract_company_name_from_text)
        if detected_company:
            normalized = normalize_company_name_for_matching(detected_company)
            result = _fuzzy_match_company_name(normalized, target_names, used_indices)
            if result:
                idx, target = result
                matched_targets[page_idx] = target
                used_indices.add(idx)

                ratio = SequenceMatcher(None, normalized, _extract_company_name_from_target(target)).ratio()
                _log_audit('company_name_match', {
                    'doc_type': doc_type,
                    'page': page_idx + 1,
                    'detected': detected_company,
                    'matched_target': target,
                    'similarity': round(ratio, 3),
                })

    unmatched_pages = [i for i, t in enumerate(matched_targets) if t is None]
    if unmatched_pages:
        remaining_targets = [(i, t) for i, t in enumerate(target_names) if i not in used_indices]
        for page_idx in unmatched_pages:
            if remaining_targets:
                fallback_idx, fallback_target = remaining_targets.pop(0)
                matched_targets[page_idx] = fallback_target
                used_indices.add(fallback_idx)
                _log(f"  [WARN] {doc_type} 第{page_idx + 1}页未识别到公司名，按顺序分配: {fallback_target}")
            else:
                page_data = pages[page_idx]
                matched_targets[page_idx] = f"未匹配_第{page_idx + 1}页.pdf"
                _log(f"  [WARN] {doc_type} 第{page_idx + 1}页无法匹配公司名")

    matched_count = sum(1 for t in matched_targets if t is not None and not t.startswith('未匹配'))
    if matched_count == 0:
        return None

    _log(f"  [OK] {doc_type} OCR 匹配: {matched_count}/{len(pages)} 页成功匹配公司名")
    return matched_targets


def export_company_named_files(input_dir: Path, output_dir: Path, target_names: List[str],
                               ocr_results: Dict[str, Dict], source_name: Optional[str], marker: str,
                               error_root: Optional[Path] = None) -> int:
    if not source_name:
        _log(f"  [ERROR] {marker} 未配置 source_pdf")
        return 0

    source_pdf = input_dir / source_name

    if not source_pdf.exists():
        _log(f"  [ERROR] {source_name} 文件不存在")
        return 0

    total_pages = inspect_pdf_page_count(source_pdf)
    doc_type = '授权书' if '授权' in marker else '所函'

    _log(f"  [INFO] {source_name}: {total_pages} 页")

    # 延迟创建错误文件夹（只在有需要核查的文件时才创建）
    error_dir: Optional[Path] = None
    def _get_error_dir() -> Path:
        nonlocal error_dir, error_root
        if error_dir is None:
            if error_root is None:
                error_root = output_dir.parent / '需人工核查'
            error_dir = error_root / doc_type
            error_dir.mkdir(parents=True, exist_ok=True)
        return error_dir

    data = _get_ocr_result(ocr_results, doc_type)
    if data and data.get('pages'):
        pages = data['pages']
        post_processor = TextPostProcessor()

        normal_ranges: List[Tuple[int, int]] = []
        normal_names: List[str] = []
        error_ranges: List[Tuple[int, int]] = []
        error_names: List[str] = []
        used_indices: set = set()

        for page_idx, page_data in enumerate(pages):
            text = page_data.get('text', '')
            detected = _extract_target_company(text, fallback_fn=post_processor.extract_company_name_from_text)
            if detected:
                normalized = normalize_company_name_for_matching(detected)
                match = _fuzzy_match_company_name(normalized, target_names, used_indices)
                if match:
                    idx, target_name = match
                    # 成功匹配，加入正常列表
                    normal_ranges.append((page_idx, page_idx + 1))
                    normal_names.append(target_name)
                    used_indices.add(idx)
                    _log(f"  [OK] {doc_type} 第{page_idx + 1}页: '{detected}' -> '{target_name}'")
                    _log_audit('company_name_match', {
                        'doc_type': doc_type,
                        'page': page_idx + 1,
                        'detected': detected,
                        'matched_target': target_name,
                    })
                else:
                    # 识别到公司名但未匹配台账，加入错误列表（台账缺失）
                    error_ranges.append((page_idx, page_idx + 1))
                    safe_detected = detected.replace('/', '_').replace('\\', '_')
                    error_name = f"{doc_type}_第{page_idx + 1}页_识别到[{safe_detected}]_台账缺失_需人工核查.pdf"
                    error_names.append(error_name)
                    _log(f"  [REVIEW] {doc_type} 第{page_idx + 1}页: '{detected}' 未在台账中找到匹配，标记为台账缺失需人工核查")
                    _log_audit('company_name_unmatched', {
                        'doc_type': doc_type,
                        'page': page_idx + 1,
                        'detected': detected,
                        'error_name': error_name,
                        'reason': '台账缺失',
                    })
            else:
                # 未识别到公司名，加入错误列表
                error_ranges.append((page_idx, page_idx + 1))
                error_name = f"{doc_type}_第{page_idx + 1}页_未识别到公司名_需人工核查.pdf"
                error_names.append(error_name)
                _log(f"  [WARN] {doc_type} 第{page_idx + 1}页: 未识别到公司名，将导出到错误文件夹")
                _log_audit('company_name_unknown', {
                    'doc_type': doc_type,
                    'page': page_idx + 1,
                    'error_name': error_name,
                })

        # 导出正常匹配的文件
        created = 0
        if normal_ranges:
            _log(f"  [INFO] {doc_type}: 台账匹配 {len(normal_ranges)} 页，导出到正常文件夹")
            created += export_pdf_ranges(source_pdf, normal_ranges, output_dir, normal_names)
        
        # 导出匹配失败的文件到错误文件夹
        if error_ranges:
            _log(f"  [INFO] {doc_type}: 匹配失败 {len(error_ranges)} 页，导出到错误文件夹")
            err_dir = _get_error_dir()
            created += export_pdf_ranges(source_pdf, error_ranges, err_dir, error_names)
        
        return created

    _log(f"  [INFO] {doc_type} 无 OCR 结果，按Excel顺序分配")
    expected_count = len(target_names)
    ranges = detect_page_ranges(total_pages, expected_count, doc_type)
    return export_pdf_ranges(source_pdf, ranges, output_dir, target_names)


def export_non_litigation_standard_outputs(sample_root: Path, input_dir: Path, output_root: Path, ocr_results: Dict[str, Dict], excel_path: Optional[Path] = None) -> Dict:
    output_root.mkdir(parents=True, exist_ok=True)
    tree = build_expected_output_tree(sample_root, excel_path=excel_path)
    created_count = 0
    _audit_log.clear()

    _log("\n[INFO] 开始导出文件...")
    _log("=" * 60)

    total_targets = sum(len(target_names) for target_names in tree.values())
    if total_targets == 0:
        _log("\n[WARN] 导出计划为空！台账中未加载到任何案件数据")
        _log(f"  请检查 Excel 文件是否存在于: {sample_root}")
        _log(f"  输入目录: {input_dir}")

    export_tasks = []
    for folder_name, target_names in tree.items():
        folder_path = output_root / folder_name
        folder_path.mkdir(parents=True, exist_ok=True)

        _log(f"\n[INFO] {folder_name} ({len(target_names)} 个文件)")

        # 统一创建错误根目录
        error_root = output_root / '需人工核查'
        error_root.mkdir(parents=True, exist_ok=True)

        if folder_name == _cfg.directory_mapping['责催']:
            count = export_notice_files(sample_root, input_dir, folder_path, ocr_results, excel_path=excel_path, error_root=error_root)
            export_tasks.append(('责催', count))

        elif folder_name == _cfg.directory_mapping['申请书']:
            count = export_application_files(input_dir, folder_path, target_names, ocr_results, error_root=error_root)
            export_tasks.append(('申请书', count))

        elif folder_name == _cfg.directory_mapping['授权书']:
            count = export_company_named_files(input_dir, folder_path, target_names, ocr_results, _cfg.doc_type_map['授权书'].source_pdf, _cfg.doc_type_map['授权书'].content_marker, error_root=error_root)
            export_tasks.append(('授权书', count))

        elif folder_name == _cfg.directory_mapping['所函']:
            count = export_company_named_files(input_dir, folder_path, target_names, ocr_results, _cfg.doc_type_map['所函'].source_pdf, _cfg.doc_type_map['所函'].content_marker, error_root=error_root)
            export_tasks.append(('所函', count))

    created_count = sum(count for _, count in export_tasks)

    if created_count == 0:
        _log("\n[WARN] 导出 0 个文件！可能原因：")
        _log(f"  1. 输入目录 ({input_dir}) 中没有 PDF 文件")
        _log(f"  2. OCR 识别未成功")
        _log(f"  3. 台账 Excel 中没有匹配的案件数据")
        _log(f"  请检查以上路径和文件是否正确")

    # 写入审计日志（仅当有问题时输出，格式为易读文本）
    if _audit_log:
        debug_dir = output_root / 'debug'
        debug_dir.mkdir(parents=True, exist_ok=True)
        audit_path = debug_dir / 'audit-log.txt'
        audit_lines = [f"OCR 识别审计报告（共 {len(_audit_log)} 条）", "=" * 60, ""]
        for i, entry in enumerate(_audit_log, 1):
            event = entry.get('event', 'unknown')
            detail = {k: v for k, v in entry.items() if k != 'event'}
            audit_lines.append(f"【{i}】事件: {event}")
            for k, v in detail.items():
                audit_lines.append(f"    {k}: {v}")
            audit_lines.append("")
        audit_path.write_text("\n".join(audit_lines), encoding='utf-8')
        _log(f"\n[INFO] 审计日志已保存: {audit_path} ({len(_audit_log)} 条)")
    # 清理旧版 json 审计日志（如存在）
    old_audit_json = output_root / 'audit-log.json'
    if old_audit_json.exists():
        old_audit_json.unlink()

    # 生成页码映射表（按 OCR 实际识别顺序）
    try:
        cases = load_non_litigation_cases(sample_root, excel_path=excel_path)
        mappings = build_company_page_mapping(cases, ocr_results)
        mapping_path = output_root / '页码映射表.xlsx'
        write_mapping_excel(cases, mappings, mapping_path)
    except Exception as e:
        _log(f"  [WARN] 生成页码映射表失败: {e}")

    # 生成需人工核查汇总 Excel
    try:
        review_summary_path = output_root / '需人工核查汇总.xlsx'
        write_review_summary_excel(_audit_log, error_root, review_summary_path)
    except Exception as e:
        _log(f"  [WARN] 生成需人工核查汇总失败: {e}")

    _log("\n" + "=" * 60)
    _log(f"[OK] 导出完成: {created_count} 个文件")

    return {
        'created_count': created_count,
        'output_root': str(output_root),
        'tree': tree,
    }


def get_actual_page_companies(ocr_results: Dict[str, Dict], doc_type: str) -> List[Optional[str]]:
    """从 OCR 结果中提取每页实际识别到的公司名（不强制匹配台账）"""
    data = _get_ocr_result(ocr_results, doc_type)
    if not data or not data.get('pages'):
        return []

    post_processor = TextPostProcessor()
    companies = []
    for page_data in data['pages']:
        text = page_data.get('text', '')
        detected = _extract_target_company(text, fallback_fn=post_processor.extract_company_name_from_text)
        companies.append(detected)
    return companies


def build_company_page_mapping(cases: List[Dict], ocr_results: Dict[str, Dict]) -> Dict[str, Dict]:
    """构建公司名到授权书/所函页码的映射（按 PDF 实际页顺序）"""
    auth_companies = get_actual_page_companies(ocr_results, '授权书')
    letter_companies = get_actual_page_companies(ocr_results, '所函')

    result = {}
    for case in cases:
        company = case['company_name']
        norm_company = normalize_company_name_for_matching(company)
        result[company] = {
            'sequence': case.get('sequence', ''),
            'notice_number': case.get('notice_number', ''),
            'auth_page': None,
            'letter_page': None,
            'auth_detected': None,
            'letter_detected': None,
        }

        # 找授权书页码
        for i, detected in enumerate(auth_companies):
            if detected and normalize_company_name_for_matching(detected) == norm_company:
                result[company]['auth_page'] = i + 1
                result[company]['auth_detected'] = detected
                break

        # 找所函页码
        for i, detected in enumerate(letter_companies):
            if detected and normalize_company_name_for_matching(detected) == norm_company:
                result[company]['letter_page'] = i + 1
                result[company]['letter_detected'] = detected
                break

    return result


def write_mapping_excel(cases: List[Dict], mappings: Dict[str, Dict], output_path: Path):
    """生成台账-页码映射 Excel，方便人工核对"""
    if not HAS_OPENPYXL:
        _log(f"  [WARN] openpyxl 未安装，跳过生成映射表")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "页码映射"

    headers = ['台账序号', '责令号', '公司名', '授权书页码', '所函页码', '授权书文件名', '所函文件名', '备注']
    ws.append(headers)

    for case in cases:
        company = case['company_name']
        mapping = mappings.get(company, {})
        auth_page = mapping.get('auth_page')
        letter_page = mapping.get('letter_page')

        auth_file = f"{company}.pdf" if auth_page else '未匹配'
        letter_file = f"{company}.pdf" if letter_page else '未匹配'

        notes = []
        if not auth_page:
            notes.append('授权书未匹配')
        if not letter_page:
            notes.append('所函未匹配')

        ws.append([
            mapping.get('sequence', case.get('sequence', '')),
            mapping.get('notice_number', case.get('notice_number', '')),
            company,
            auth_page if auth_page else '未匹配',
            letter_page if letter_page else '未匹配',
            auth_file,
            letter_file,
            '；'.join(notes) if notes else '',
        ])

    wb.save(output_path)
    _log(f"  [OK] 页码映射表已保存: {output_path}")


def write_review_summary_excel(audit_log: List[Dict], error_root: Optional[Path], output_path: Path):
    """生成需人工核查汇总 Excel，方便用户一次性查看所有需要人工处理的文件"""
    try:
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False

    if not HAS_OPENPYXL:
        _log(f"  [WARN] openpyxl 未安装，跳过生成需人工核查汇总")
        return

    # 筛选需要人工核查的事件
    review_events = []
    for entry in audit_log:
        event = entry.get('event', '')
        if event in {'application_unknown', 'application_unmatched', 'application_needs_review',
                     'company_name_unmatched', 'company_name_unknown'}:
            review_events.append(entry)

    if not review_events:
        # 即使无需核查也生成空表，让用户确认全部通过
        wb = Workbook()
        ws = wb.active
        ws.title = "需人工核查汇总"
        try:
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            ws.merge_cells('A1:H1')
            title_cell = ws['A1']
            title_cell.value = '需人工核查汇总（全部通过，无需核查）'
            title_cell.font = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')
            title_cell.fill = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 35
            headers = ['序号', '文档类型', '文件名', '页码范围', '识别到的内容', '问题描述', '建议操作', '台账匹配结果']
            ws.append(headers)
            for col in range(1, 9):
                cell = ws.cell(row=2, column=col)
                cell.font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
                cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.merge_cells('A3:H3')
            pass_cell = ws['A3']
            pass_cell.value = '所有文件识别结果均已精确匹配，无需人工核查'
            pass_cell.font = Font(name='微软雅黑', size=12, color='548235')
            pass_cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[3].height = 30
            wb.save(output_path)
            _log(f"  [OK] 需人工核查汇总已保存（全部通过）: {output_path}")
        except Exception as e:
            _log(f"  [WARN] 生成空核查汇总失败: {e}")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "需人工核查汇总"

    # 样式定义
    title_font = Font(name='微软雅黑', bold=True, size=14, color='FFFFFF')
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    body_font = Font(name='微软雅黑', size=10)
    body_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    title_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    thin_border = Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )

    # 文档类型颜色（奇偶行交替）
    type_colors = {
        '申请书': ('FFF2CC', 'FFFDE7'),
        '授权书': ('D4E6F1', 'E8F4FC'),
        '所函': ('D5E8D4', 'E8F8ED'),
    }

    # 设置列宽
    col_widths = [6, 10, 55, 12, 40, 30, 35, 25]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # 第1行：标题
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = f'需人工核查汇总（共 {len(review_events)} 项）'
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = center_alignment
    title_cell.border = thin_border
    ws.row_dimensions[1].height = 35

    # 第2行：表头
    headers = ['序号', '文档类型', '文件名', '页码范围', '识别到的内容', '问题描述', '建议操作', '台账匹配结果']
    ws.append(headers)
    for col in range(1, 9):
        cell = ws.cell(row=2, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行从第3行开始
    for idx, entry in enumerate(review_events, 1):
        row_num = idx + 2  # 数据行号：3, 4, 5...
        event = entry.get('event', '')
        doc_type = ''
        detected = ''
        pages = entry.get('pages', '')
        matched_target = entry.get('matched_target', '')
        reason = ''
        suggestion = ''
        filename = ''

        if event.startswith('application_'):
            doc_type = '申请书'
            detected = entry.get('detected_notice', '')
            filename = f"申请书_{pages}_需人工核查.pdf"
            if event == 'application_unknown':
                suggestion = '请人工确认责令号并匹配到正确的台账案件'
                reason = '未识别到责令号'
            elif event == 'application_unmatched':
                suggestion = 'OCR识别的责令号与台账不完全匹配，请核对并手动修正'
                reason = '识别到的责令号未匹配台账'
            elif event == 'application_needs_review':
                suggestion = '存在模糊匹配，请人工确认是否正确'
                reason = '模糊匹配需人工确认'

        elif event.startswith('company_name_'):
            doc_type = entry.get('doc_type', '')
            detected = entry.get('detected', '')
            page = entry.get('page', '')
            filename = f"{doc_type}_第{page}页_需人工核查.pdf"
            if event == 'company_name_unmatched':
                suggestion = '识别的公司名与台账不完全匹配，请核对并手动修正'
                reason = '公司名未匹配台账'
            elif event == 'company_name_unknown':
                suggestion = '未能识别到公司名，请人工补充'
                reason = '未识别到公司名'

        row_data = [
            idx,
            doc_type,
            filename,
            pages if pages else f"第{page}页" if page else '',
            detected,
            reason,
            suggestion,
            matched_target if matched_target else '未匹配到台账',
        ]
        ws.append(row_data)

        # 设置行样式
        colors = type_colors.get(doc_type, ('F5F5F5', 'FFFFFF'))
        fill_color = colors[idx % 2]
        row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col in range(1, 9):
            cell = ws.cell(row=row_num, column=col)
            cell.font = body_font
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = body_alignment if col > 1 else center_alignment

        # 页码列居中
        ws.cell(row=row_num, column=4).alignment = center_alignment
        # 文档类型加粗
        ws.cell(row=row_num, column=2).font = Font(name='微软雅黑', size=10, bold=True)

    wb.save(output_path)
    _log(f"  [OK] 需人工核查汇总已保存: {output_path} ({len(review_events)} 条)")
