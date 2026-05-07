#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from typing import Dict

from openpyxl import load_workbook

from rename_rules import build_notice_rename_plan, RenameService


def normalize_notice_number(text: str) -> str:
    value = str(text).strip()
    value = value.replace('（', '〔').replace('）', '〕').replace('(', '〔').replace(')', '〕')
    return value


def load_ledger_map_from_excel(excel_path: Path) -> Dict[str, str]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    ledger_map: Dict[str, str] = {}

    for row in sheet.iter_rows(values_only=True):
        values = [str(item).strip() if item is not None else '' for item in row]
        if len(values) < 3:
            continue
        original_notice = values[1]
        renamed_notice = values[2]
        if '责字' not in original_notice or '责催' not in renamed_notice:
            continue

        normalized_notice = normalize_notice_number(original_notice)
        sequence_match = __import__('re').match(r'(\d+)-责催-', renamed_notice.replace(' ', ''))
        if not sequence_match:
            continue
        ledger_map[normalized_notice] = sequence_match.group(1)

    return ledger_map


def build_notice_plan_from_paths(input_dir: Path, output_dir: Path, ledger_map: Dict[str, str], service: RenameService) -> Dict:
    samples = []
    for pdf_path in sorted(input_dir.glob('*.pdf')):
        text_path = output_dir / f'{pdf_path.stem}_ultra_result.txt'
        if text_path.exists():
            samples.append({
                'filename': pdf_path.name,
                'ocr_text': text_path.read_text(encoding='utf-8'),
            })
        else:
            samples.append({
                'filename': pdf_path.name,
                'ocr_text': '',
            })
    return build_notice_rename_plan(samples=samples, ledger_map=ledger_map, service=service)


def summarize_plan(plan: Dict) -> Dict:
    matched_items = [item for item in plan['items'] if item['matched']]
    unmatched_items = [item for item in plan['items'] if not item['matched']]
    return {
        'total': plan['total'],
        'matched': plan['matched'],
        'unmatched': plan['unmatched'],
        'matched_items': matched_items,
        'unmatched_items': unmatched_items,
    }
