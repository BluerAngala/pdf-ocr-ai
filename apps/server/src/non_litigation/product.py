#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import shutil
import tempfile
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from core.config_loader import load_config

_cfg = load_config()


def _safe_load_workbook(path):
    if getattr(os, "frozen", False):
        tmp_dir = tempfile.mkdtemp()
        tmp_path = os.path.join(tmp_dir, "workbook.xlsx")
        shutil.copy2(str(path), tmp_path)
        try:
            return load_workbook(tmp_path, data_only=True)
        finally:
            try:
                os.remove(tmp_path)
                os.rmdir(tmp_dir)
            except OSError:
                pass
    return load_workbook(path, data_only=True)


def _extract_sequence_from_renamed(renamed_notice: str) -> str:
    """
    从 "1-责催-..." / "1-授权书-..." / "1-申请书pdf-..." 提取序列号。
    返回纯数字字符串。
    """
    # 在常见命名分隔符前切分，取首段再提取数字
    for sep in ('-责催-', '-授权书-', '-申请书pdf-', '-申请书-', '-所函-'):
        if sep in renamed_notice:
            head = renamed_notice.split(sep, 1)[0].strip()
            digits = re.findall(r'\d+', head)
            if digits:
                return digits[0]
    # 兜底：取首段连续数字
    digits = re.findall(r'\d+', renamed_notice)
    return digits[0] if digits else ''


def _matches_any_keyword(text: str, keyword_pattern: str) -> bool:
    """
    判断 text 是否命中 keyword_pattern 中的任一关键字。
    keyword_pattern 支持 "|" 分隔的多关键字（如 "责催-|授权书-|申请书pdf-"）。
    """
    if not keyword_pattern:
        return False
    for kw in keyword_pattern.split('|'):
        if kw and kw in text:
            return True
    return False


def load_non_litigation_cases(sample_root: Path, excel_path: Optional[Path] = None) -> List[Dict]:
    resolved = excel_path if excel_path else sample_root / _cfg.excel_filename
    workbook = _safe_load_workbook(resolved)
    sheet = workbook.active
    cases: List[Dict] = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(item).strip() if item is not None else '' for item in row]
        if len(values) < _cfg.excel_min_columns:
            continue
        original_notice = values[_cfg.excel_column_original_notice]
        renamed_notice = values[_cfg.excel_column_renamed_notice]
        company_name = values[_cfg.excel_column_company_name]
        if (
            _cfg.excel_filter_original_notice not in original_notice
            or not _matches_any_keyword(renamed_notice, _cfg.excel_filter_renamed_notice)
            or not company_name
        ):
            continue
        sequence = _extract_sequence_from_renamed(renamed_notice)
        notice_number = original_notice.replace(' ', '')
        if not sequence or not notice_number:
            continue
        cases.append({
            'sequence': sequence,
            'notice_number': notice_number,
            'company_name': company_name,
        })
    return cases


def build_non_litigation_standard_plan(sample_root: Path, excel_path: Optional[Path] = None) -> Dict[str, List[Dict]]:
    cases = load_non_litigation_cases(sample_root, excel_path=excel_path)
    plan = {dt.key: [] for dt in _cfg.doc_types}
    for case in cases:
        sequence = case['sequence']
        notice_number = case['notice_number']
        company_name = case['company_name']
        plan[_cfg.notice_doc_type.key].append({
            'target_filename': _cfg.notice_doc_type.filename_pattern.format(sequence=sequence, notice_number=notice_number),
            'company_name': company_name,
        })
        plan[_cfg.doc_type_map['申请书'].key].append({
            'target_filename': _cfg.doc_type_map['申请书'].filename_pattern.format(sequence=sequence, notice_number=notice_number),
            'company_name': company_name,
        })
        plan[_cfg.doc_type_map['授权书'].key].append({
            'target_filename': _cfg.doc_type_map['授权书'].filename_pattern.format(company_name=company_name),
            'company_name': company_name,
        })
        plan[_cfg.doc_type_map['所函'].key].append({
            'target_filename': _cfg.doc_type_map['所函'].filename_pattern.format(company_name=company_name),
            'company_name': company_name,
        })
    return plan
