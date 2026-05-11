#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
强制执行组 - 裁定信息提取与导出模块

功能：
1. 批量处理裁定PDF，提取信息
2. 与台账（非诉表格）进行责令号匹配
3. 导出合并后Excel（原表字段 + OCR识别字段）
"""

import json
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Any
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from paths import ROOT

from config_loader import load_config
from enforcement_extractor import RulingPDFExtractor, RulingInfo, extract_ruling_from_pdf, chinese_date_to_arabic
from enforcement_product import EnforcementCaseRegistry, load_enforcement_cases

_cfg = load_config()
_enforcement_cfg = _cfg.raw_config.get('enforcement', {})
_paths_cfg = _enforcement_cfg.get('paths', {})


def build_output_excel(
    registry: EnforcementCaseRegistry,
    pdf_results: Dict[str, RulingInfo],
    output_path: Path,
) -> Path:
    """
    构建输出Excel：合并非诉表格字段与OCR识别字段

    输出列：
      区号 | 行政审查案号 | 责令号 | 被执行人 | 职工姓名 | 金额 | 法官/法官助理 | 执行时间 | 裁定结果 | 备注
    """
    rows = []
    for case in registry.cases:
        matched_info = _match_case_to_pdf(case, pdf_results)

        row = {
            '区号': case.region or '',
            '行政审查案号': '',
            '责令号': case.notice_number or '',
            '被执行人': case.respondent or '',
            '职工姓名': case.employee or '',
            '金额': _format_amount(case.amount),
            '法官/法官助理': '',
            '执行时间': '',
            '裁定结果': '',
            '备注': '',
        }

        if matched_info:
            row['行政审查案号'] = matched_info.court_case_number or ''
            row['执行时间'] = _format_date(matched_info.ruling_date) or ''
            row['裁定结果'] = matched_info.ruling_result or ''
            judge_parts = []
            if matched_info.judge:
                judge_parts.append(matched_info.judge)
            if matched_info.clerk:
                judge_parts.append(matched_info.clerk)
            row['法官/法官助理'] = '/'.join(judge_parts) if judge_parts else ''

        rows.append(row)

    _sort_rows_by_region(rows)
    df = pd.DataFrame(rows)
    df = df.fillna('')
    columns_order = ['区号', '行政审查案号', '责令号', '被执行人', '职工姓名', '金额', '法官/法官助理', '执行时间', '裁定结果', '备注']
    df = df[columns_order]

    with pd.ExcelWriter(str(output_path), engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='执行组识别结果')
        ws = writer.sheets['执行组识别结果']
        _auto_fit_columns(ws)
        _apply_header_style(ws)

    return output_path


def _match_case_to_pdf(case, pdf_results: Dict[str, RulingInfo]) -> Optional[RulingInfo]:
    for info in pdf_results.values():
        for ocr_notice in info.notice_numbers:
            norm_ocr = _normalize_for_match(ocr_notice)
            norm_excel = _normalize_for_match(case.notice_number)
            if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                return info
    return None


def _normalize_for_match(text: str) -> str:
    """标准化责令号用于匹配：统一括号、去空格"""
    if not text:
        return ''
    text = str(text).replace(' ', '')
    for old, new in [('(', '〔'), (')', '〕'), ('（', '〔'), ('）', '〕'),
                      ('[', '〔'), (']', '〕'), ('［', '〔'), ('］', '〕'),
                      ('【', '〔'), ('】', '〕')]:
        text = text.replace(old, new)
    return text


def _format_amount(amount) -> str:
    if amount is None:
        return ''
    return str(int(amount)) if amount == int(amount) else str(amount)


def _format_date(date_str: Optional[str]) -> str:
    if not date_str:
        return ''
    try:
        return chinese_date_to_arabic(date_str)
    except Exception:
        return date_str


def _sort_rows_by_region(rows: List[Dict]):
    region_order = {'萝岗': 1, '越秀': 2, '荔湾': 3, '天河': 4, '海珠': 5,
                    '白云': 6, '黄埔': 7, '番禺': 8, '花都': 9, '南沙': 10,
                    '增城': 11, '从化': 12}
    rows.sort(key=lambda r: (region_order.get(r.get('区号', ''), 99), r.get('责令号', '')))


def _auto_fit_columns(ws):
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                cell_len = 0
                for ch in str(cell.value):
                    cell_len += 2 if '\u4e00' <= ch <= '\u9fff' else 1
                max_length = max(max_length, cell_len)
        adjusted = min(max_length + 4, 60)
        ws.column_dimensions[col_letter].width = max(adjusted, 8)


def _apply_header_style(ws):
    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(name='微软雅黑', bold=True, size=11)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment


def run_enforcement_extraction(
    input_dir: Path,
    excel_path: Path,
    output_dir: Path,
    use_ocr: bool = False,
) -> Dict[str, Any]:
    """
    运行强制执行组完整提取+导出流程

    流程：
    1. 加载台账（非诉表格.xlsx）
    2. 批量OCR识别裁定PDF（行审案号、责令号、法官/法官助理、责令作出时间）
    3. 按责令号匹配台账与OCR结果
    4. 导出合并后Excel
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    print("=" * 60)
    print("强制执行组 - 裁定信息识别与导出")
    print("=" * 60)

    print("\n[1/4] 加载台账数据...")
    registry = load_enforcement_cases(excel_path)

    print("\n[2/4] 识别裁定PDF...")
    pdf_results = {}
    extractor = RulingPDFExtractor(use_ocr=use_ocr)
    pdf_files = sorted(input_dir.glob("*.pdf"))
    print(f"  发现 {len(pdf_files)} 个PDF文件")

    for pdf_file in pdf_files:
        print(f"  处理: {pdf_file.name}")
        try:
            info = extractor.extract_from_pdf(pdf_file)
            key = info.court_case_number if info.court_case_number else pdf_file.stem
            pdf_results[key] = info
            print(f"    案号: {info.court_case_number}")
            print(f"    责令号: {info.notice_numbers}")
            print(f"    法官: {info.judge}, 法官助理/书记员: {info.clerk}")
            print(f"    裁定日期: {info.ruling_date}")
            print(f"    裁定结果: {info.ruling_result}")
        except Exception as e:
            print(f"    [ERROR] 处理失败: {e}")

    print("\n[3/4] 匹配责令号并合并数据...")
    stats = _match_and_stats(registry, pdf_results)

    print("\n[4/4] 导出结果...")

    excel_output_path = output_dir / _paths_cfg.get('excel_output_filename', '执行组识别结果.xlsx')
    build_output_excel(registry, pdf_results, excel_output_path)

    json_path = output_dir / _paths_cfg.get('json_output_filename', 'enforcement_extracted.json')
    _export_json(pdf_results, stats, json_path)

    summary_path = output_dir / 'enforcement_summary.json'
    _export_summary(registry, pdf_results, summary_path)

    print("\n" + "=" * 60)
    print("处理统计")
    print("=" * 60)
    print(f"  PDF文件数: {stats['total_pdfs']}")
    print(f"  台账行数: {stats['total_excel_rows']}")
    print(f"  成功匹配: {stats['matched_rows']}")
    print(f"  未匹配: {stats['unmatched_rows']}")
    print(f"  撤回执行: {stats['withdraw_count']}")
    print(f"\n输出文件:")
    print(f"  - {excel_output_path}")
    print(f"  - {json_path}")
    print(f"  - {summary_path}")

    return {
        'stats': stats,
        'output_files': {
            'excel': str(excel_output_path),
            'json': str(json_path),
            'summary': str(summary_path),
        }
    }


def _match_and_stats(registry: EnforcementCaseRegistry, pdf_results: Dict[str, RulingInfo]) -> Dict[str, Any]:
    matched = set()
    withdraw_count = sum(1 for info in pdf_results.values() if info.is_withdraw)
    for case in registry.cases:
        for info in pdf_results.values():
            for ocr_notice in info.notice_numbers:
                norm_ocr = _normalize_for_match(ocr_notice)
                norm_excel = _normalize_for_match(case.notice_number)
                if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                    matched.add(case.notice_number)
                    break

    return {
        'total_pdfs': len(pdf_results),
        'total_excel_rows': len(registry.cases),
        'matched_rows': len(matched),
        'unmatched_rows': len(registry.cases) - len(matched),
        'withdraw_count': withdraw_count,
    }


def _export_json(pdf_results: Dict[str, RulingInfo], stats: Dict, output_path: Path):
    data = {
        'export_time': datetime.now().isoformat(),
        'stats': stats,
        'pdf_results': {k: v.to_dict() for k, v in pdf_results.items()},
    }
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"  JSON导出: {output_path}")


def _export_summary(registry: EnforcementCaseRegistry, pdf_results: Dict[str, RulingInfo], output_path: Path):
    summary = {
        'export_time': datetime.now().isoformat(),
        'excel_rows': len(registry.cases),
        'pdf_files': len(pdf_results),
        'matched': [],
        'unmatched_excel': [],
        'unmatched_pdfs': [],
    }

    matched_excel = set()
    for case in registry.cases:
        found = False
        for info in pdf_results.values():
            for ocr_notice in info.notice_numbers:
                norm_ocr = _normalize_for_match(ocr_notice)
                norm_excel = _normalize_for_match(case.notice_number)
                if norm_ocr.endswith(norm_excel) or norm_excel.endswith(norm_ocr):
                    found = True
                    matched_excel.add(case.notice_number)
                    summary['matched'].append({
                        '责令号': case.notice_number,
                        '职工': case.employee,
                        '行审案号': info.court_case_number,
                        '法官': info.judge,
                        '日期': info.ruling_date,
                    })
                    break
            if found:
                break
        if not found:
            summary['unmatched_excel'].append({
                '责令号': case.notice_number,
                '被执行人': case.respondent,
                '职工': case.employee,
            })

    for key, info in pdf_results.items():
        if not any(m['行审案号'] == info.court_case_number for m in summary['matched']):
            summary['unmatched_pdfs'].append({
                '文件': key,
                '行审案号': info.court_case_number,
                '责令号': info.notice_numbers,
            })

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)
    print(f"  汇总JSON导出: {output_path}")


if __name__ == "__main__":
    input_dir = Path("样本材料/强制组-自动化/提取信息")
    excel_path = Path("样本材料/强制组-自动化/提取信息/非诉表格.xlsx")
    output_dir = Path("output/enforcement")

    if input_dir.exists() and excel_path.exists():
        result = run_enforcement_extraction(input_dir, excel_path, output_dir, use_ocr=False)
        print("\n[OK] 处理完成!")
    else:
        print(f"[ERROR] 输入路径不存在")
        print(f"  输入目录: {input_dir} (存在: {input_dir.exists()})")
        print(f"  Excel文件: {excel_path} (存在: {excel_path.exists()})")
