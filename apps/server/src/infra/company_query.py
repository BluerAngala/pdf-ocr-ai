#!/usr/bin/env python3

import json
import threading
import time
from collections import OrderedDict
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
import requests

from core.config_loader import _load_config
from core.paths import USER_DATA_DIR


class CompanyQueryError(Exception):
    def __init__(self, message: str, raw_response: dict = None, balance_depleted: bool = False, recharge_url: str = None):
        self.message = message
        self.raw_response = raw_response
        self.balance_depleted = balance_depleted
        self.recharge_url = recharge_url
        super().__init__(self.message)


from core.task_cancel import request_cancel, is_cancelled, clear as clear_cancel


def _wait_with_timeout(futures, timeout=0.5):
    fs = set(futures.keys()) if isinstance(futures, dict) else futures
    if not fs:
        return set(), set()
    done, not_done = wait(fs, timeout=timeout, return_when=FIRST_COMPLETED)
    return done, not_done


def _get_cache_path(excel_path: Path) -> Path:
    output_dir = USER_DATA_DIR / "output" / "company-query"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir / f"company_query_cache_{excel_path.stem}.json"


def _load_cache(cache_path: Path) -> Dict[str, dict]:
    if not cache_path.exists():
        return {}
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {item["original_name"]: item for item in data.get("results", []) if "original_name" in item}
    except Exception:
        return {}


def _save_cache(cache_path: Path, results: List[dict]):
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump({"results": results, "updated_at": datetime.now().isoformat()}, f, ensure_ascii=False, indent=2)


def _is_cache_expired(cache_path: Path, ttl_days: int) -> bool:
    if ttl_days <= 0:
        return True
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        updated_at = data.get("updated_at", "")
        if not updated_at:
            return True
        updated_time = datetime.fromisoformat(updated_at)
        return (datetime.now() - updated_time).days >= ttl_days
    except Exception:
        return True


def load_cached_results(excel_path: str, ttl_days: int = 0) -> List[dict]:
    cache_path = _get_cache_path(Path(excel_path))
    if not cache_path.exists():
        return []
    if _is_cache_expired(cache_path, ttl_days):
        return []
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("results", [])
    except Exception:
        return []


def clear_cache(excel_path: str) -> bool:
    """清除指定 Excel 文件的缓存"""
    cache_path = _get_cache_path(Path(excel_path))
    if cache_path.exists():
        try:
            cache_path.unlink()
            return True
        except Exception:
            return False
    return True  # 缓存文件不存在，视为清除成功


def _write_results_excel(results: List[dict], output_path: Path, source_df: pd.DataFrame = None, start_idx: int = 0):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    if source_df is not None:
        result_df = source_df.iloc[start_idx:start_idx + len(results)].copy()
    else:
        names = [r.get("original_name", "") for r in results]
        result_df = pd.DataFrame({"被执行人": names})

    for col, key in [("现用名", "current_name"), ("法代", "legal_person"), ("所在地", "location"), ("社会信用代码", "credit_code")]:
        if col not in result_df.columns:
            result_df[col] = ""
        values = [r.get(key, "") for r in results]
        if len(values) < len(result_df):
            values += [""] * (len(result_df) - len(values))
        result_df[col] = values[:len(result_df)]

    col_order = []
    for c in ["序号", "序号.", "No.", "no.", "NO.", "Index", "index"]:
        if c in result_df.columns:
            col_order.append(c)
            break
    for c in ["被执行人", "现用名", "法代", "所在地", "社会信用代码"]:
        if c in result_df.columns:
            col_order.append(c)
    for c in result_df.columns:
        if c not in col_order:
            col_order.append(c)
    result_df = result_df[col_order]

    wb = Workbook()
    ws = wb.active
    ws.title = "企业查询结果"

    headers = list(result_df.columns)
    ws.append(headers)

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B4C6E7"),
        right=Side(style="thin", color="B4C6E7"),
        top=Side(style="thin", color="B4C6E7"),
        bottom=Side(style="thin", color="B4C6E7"),
    )
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    data_font = Font(name="微软雅黑", size=10)
    data_align = Alignment(vertical="center", wrap_text=True)
    success_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    failed_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    warning_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")

    for row_idx, (_, row_data) in enumerate(result_df.iterrows(), 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header, ""))
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

        status = results[row_idx - 2].get("status", "") if row_idx - 2 < len(results) else ""
        fill = None
        if status == "success":
            fill = success_fill
        elif status == "failed":
            fill = failed_fill
        elif status == "warning":
            fill = warning_fill
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    col_widths = {"被执行人": 28, "现用名": 28, "法代": 14, "所在地": 30, "社会信用代码": 22}
    for col_idx, header in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(header, 15)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return str(output_path.resolve())


def export_cache_to_excel(excel_path: str) -> dict:
    cache_path = _get_cache_path(Path(excel_path))
    if not cache_path.exists():
        return {"exported": False, "error": "缓存文件不存在，请先执行一次查询"}

    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        results = data.get("results", [])
        if not results:
            return {"exported": False, "error": "缓存为空，请先执行一次查询"}

        try:
            df = pd.read_excel(excel_path, dtype=str)
        except Exception:
            df = None

        from core.paths import USER_DATA_DIR
        output_dir = USER_DATA_DIR / "output" / "company-query"
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"企业查询缓存_{Path(excel_path).stem}_{timestamp}.xlsx"

        output_excel_path = _write_results_excel(results, output_path, source_df=df)

        return {
            "exported": True,
            "output_excel_path": output_excel_path,
            "total": len(results),
            "success_count": sum(1 for r in results if r.get("status") == "success"),
            "warning_count": sum(1 for r in results if r.get("status") == "warning"),
            "fail_count": sum(1 for r in results if r.get("status") == "failed"),
        }
    except Exception as e:
        return {"exported": False, "error": str(e)}


def _get_api_config() -> dict:
    raw = _load_config()
    cq = raw.get("company_query", {})
    return {
        "api_url": cq.get("api_url", ""),
        "userid": cq.get("userid", ""),
        "userkey": cq.get("userkey", ""),
        "request_delay": cq.get("request_delay", 0.5),
        "max_concurrent": cq.get("max_concurrent", 3),
        "excel_column_name": cq.get("excel_column_name", "被执行人"),
    }


def _api_call(config: dict, method: str, params: dict) -> dict:
    url = f"{config['api_url'].rstrip('/')}/{method}"
    response = requests.post(url, json=params, timeout=60)
    try:
        data = response.json()
        if isinstance(data, dict) and "code" in data:
            return data
        if isinstance(data, dict) and "errDetail" in data:
            detail = data.get('errDetail', '')
            err_msg = data.get('errMsg', '')
            return {"code": 500, "msg": f"服务端异常: {err_msg} | {detail}"[:400]}
    except Exception:
        pass
    if response.status_code >= 500:
        return {"code": response.status_code, "msg": f"HTTP {response.status_code}: {response.text[:200]}"}
    response.raise_for_status()
    return {"code": 500, "msg": f"非预期响应 (HTTP {response.status_code})"}


def check_account() -> dict:
    config = _get_api_config()
    userid = config.get("userid", "")
    userkey = config.get("userkey", "")
    api_url = config.get("api_url", "")

    if not userid or not userkey or not api_url:
        return {"status": "error", "userid": userid, "message": "API 配置缺失（userid/userkey/api_url 为空）"}

    try:
        result = _api_call(config, "getBalance", {"userid": userid, "userkey": userkey})
        code = result.get("code", 0)
        msg = result.get("msg", "")
        data = result.get("data", {}) or {}

        if code == 200:
            remaining = data.get("remainingTimes", 0)
            base = {
                "userid": userid,
                "message": msg or "账号正常",
                "usedTimes": data.get("usedTimes", 0),
                "remainingTimes": remaining,
                "totalLimit": data.get("totalLimit", 0),
                "userName": data.get("userName", ""),
            }
            if remaining <= 0:
                base["status"] = "depleted"
                base["message"] = "余额不足，请充值"
            else:
                base["status"] = "ok"
            return base

        if code == 400 and "余额不足" in msg:
            return {
                "status": "depleted",
                "userid": userid,
                "message": msg,
                "usedTimes": data.get("usedTimes", 0),
                "remainingTimes": data.get("remainingTimes", 0),
                "totalLimit": data.get("totalLimit", 0),
            }
        return {"status": "error", "userid": userid, "message": f"API 返回错误: {msg}"}
    except requests.Timeout:
        return {"status": "error", "userid": userid, "message": "API 连接超时"}
    except requests.ConnectionError:
        return {"status": "error", "userid": userid, "message": "API 连接失败，请检查网络"}
    except Exception as e:
        return {"status": "error", "userid": userid, "message": f"检查失败: {e}"}


def recharge(code: str) -> dict:
    config = _get_api_config()
    userid = config.get("userid", "")
    userkey = config.get("userkey", "")
    if not userid or not userkey:
        return {"success": False, "message": "API 配置缺失"}
    if not code or not code.strip():
        return {"success": False, "message": "兑换码不能为空"}
    try:
        result = _api_call(config, "recharge", {"userid": userid, "userkey": userkey, "code": code.strip()})
        rcode = result.get("code", 0)
        msg = result.get("msg", "")
        data = result.get("data", {}) or {}
        if rcode == 200:
            return {
                "success": True,
                "message": msg,
                "addTimes": data.get("addTimes", 0),
                "afterRemaining": data.get("afterRemaining", 0),
                "totalLimit": data.get("totalLimit", 0),
            }
        return {"success": False, "message": msg}
    except Exception as e:
        return {"success": False, "message": f"充值失败: {e}"}


def query_company_info(company_name: str, config: dict) -> dict:
    return _api_call(config, "search", {
        "userid": config["userid"],
        "userkey": config["userkey"],
        "companyName": company_name,
    })

def get_company_data(company_name: str, config: dict) -> dict:
    result = query_company_info(company_name, config)
    code = result.get("code", 0)
    msg = result.get("msg", "")
    data = result.get("data", {}) or {}

    if code == 400 and "余额不足" in msg:
        raise CompanyQueryError(
            f"API 余额不足: {msg}",
            raw_response=result,
            balance_depleted=True,
            recharge_url=data.get("rechargeUrl", ""),
        )
    if code != 200:
        raise CompanyQueryError(
            f"API 返回错误: {msg}",
            raw_response=result,
        )
    data = result.get("data", {})
    company_info = data.get("companyInfo")
    if not company_info:
        raise CompanyQueryError("未查询到企业数据", raw_response=result)

    inner_data = company_info.get("data")
    if isinstance(inner_data, str):
        try:
            inner_data = json.loads(inner_data)
        except (json.JSONDecodeError, TypeError):
            raise CompanyQueryError("企业数据解析失败", raw_response=result)
        company_data = inner_data.get("data", {})
    elif isinstance(inner_data, dict):
        company_data = inner_data.get("data", {})
    else:
        company_data = company_info

    if not company_data:
        raise CompanyQueryError("未查询到企业数据", raw_response=result)
    return company_data


def format_current_name(company_data: dict) -> str:
    current_name = company_data.get("CompanyName", "")
    history_names = company_data.get("HistoryNames", "")
    if history_names and history_names != current_name:
        return f"{current_name}（曾用名：{history_names}）"
    return current_name


def extract_location(company_data: dict) -> str:
    """提取企业住所地（完整地址），优先使用详细地址字段"""
    # 1. 尝试多种可能包含完整地址的字段（CompanyAddress 是 API 实际返回的字段，优先级最高）
    addr_keys = [
        "CompanyAddress", "Address", "RegAddress", "OfficeAddress",
        "Domicile", "RegisteredAddress", "DetailAddress",
        "住所", "住所地", "注册地址", "经营地址", "地址",
    ]
    for key in addr_keys:
        val = company_data.get(key, "")
        if val and len(val) > 4:  # 地址字段有值即可返回
            return val.strip()

    # 2. 回退拼接 Province + City + District
    province = company_data.get("Province", "") or company_data.get("province", "")
    city = company_data.get("City", "") or company_data.get("city", "")
    district = company_data.get("District", "") or company_data.get("Area", "") or company_data.get("district", "")
    parts = [p for p in [province, city, district] if p]
    return "".join(parts) if parts else ""


def _validate_company_result(result: dict) -> tuple[str, str]:
    """
    验证企业查询结果
    返回: (status, error_message)
    - 所有字段都为空 -> failed
    - 部分字段为空 -> warning
    - 所有字段都有值 -> success
    """
    fields = ["current_name", "legal_person", "location", "credit_code"]
    empty_fields = []
    
    for field in fields:
        if not result.get(field):
            empty_fields.append(field)
    
    if len(empty_fields) == len(fields):
        return "failed", "查询结果所有字段均为空"
    elif empty_fields:
        field_names = {
            "current_name": "现用名",
            "legal_person": "法定代表人", 
            "location": "所在地",
            "credit_code": "信用代码"
        }
        missing = "、".join([field_names.get(f, f) for f in empty_fields])
        return "warning", f"缺少字段: {missing}"
    else:
        return "success", ""


def process_single_company(company_name: str, config: dict) -> dict:
    try:
        company_data = get_company_data(company_name, config)
        location = extract_location(company_data)
        result = {
            "original_name": company_name,
            "current_name": format_current_name(company_data),
            "legal_person": company_data.get("LegalPerson", ""),
            "location": location,
            "credit_code": company_data.get("CreditNo", ""),
        }
        
        # 如果location较短（可能只有区），尝试从原始数据中找更完整的地址
        if len(location) <= 6:
            # 记录所有含地址信息的字段，优先取最长的
            addr_candidates = {}
            for k, v in company_data.items():
                if isinstance(v, str) and v and any(
                    kw in k for kw in ["addr", "Addr", "地址", "住所", "位置", "loc", "Loc", "domicile", "Domicile"]
                ):
                    addr_candidates[k] = v
            if addr_candidates:
                best = max(addr_candidates.values(), key=len)
                if len(best) > len(location):
                    result["location"] = best
            else:
                # 没找到地址字段，记录可用字段名方便排查
                import logging
                logging.getLogger(__name__).debug(
                    "企业查询地址字段不足: company=%s, location=%s, 可用字段=%s",
                    company_name, location, list(company_data.keys())
                )
        
        status, error_msg = _validate_company_result(result)
        result["status"] = status
        if error_msg:
            result["error"] = error_msg
            
        return result
        
    except CompanyQueryError as e:
        result = {
            "original_name": company_name,
            "current_name": "",
            "legal_person": "",
            "location": "",
            "credit_code": "",
            "status": "failed",
            "error": e.message,
            "balance_depleted": e.balance_depleted,
            "recharge_url": e.recharge_url,
        }
        return result
    except Exception as e:
        return {
            "original_name": company_name,
            "current_name": "",
            "legal_person": "",
            "location": "",
            "credit_code": "",
            "status": "failed",
            "error": str(e),
        }


def process_company_query(
    excel_path: Path,
    range_start: int = 1,
    range_end: int = 99999,
    cache_ttl_days: int = 0,
    task_id: str = "",
    emitter=None,
    output_dir: Path = None,
) -> dict:
    config = _get_api_config()
    column_name = config["excel_column_name"]
    cache_path = _get_cache_path(excel_path)

    clear_cancel(task_id)

    df = pd.read_excel(excel_path, dtype=str)
    if column_name not in df.columns:
        raise ValueError(f"Excel 中缺少 '{column_name}' 列，可用列: {list(df.columns)}")

    all_names = df[column_name].dropna().tolist()

    start_idx = max(0, range_start - 1)
    end_idx = len(all_names) if range_end >= 99999 else min(range_end, len(all_names))
    company_names = all_names[start_idx:end_idx]

    total = len(company_names)

    raw_cache = _load_cache(cache_path)
    cache: Dict[str, dict] = {}
    if cache_ttl_days > 0 and not _is_cache_expired(cache_path, cache_ttl_days):
        cache = raw_cache

    results: OrderedDict[str, dict] = OrderedDict()
    skipped = 0
    completed_count = 0
    balance_depleted = False
    recharge_url = ""

    cached_names = []
    uncached_names = []

    for name in company_names:
        if name in cache:
            cached_names.append(name)
        else:
            uncached_names.append(name)

    for name in cached_names:
        if is_cancelled(task_id):
            if emitter:
                emitter.log("warn", f"任务已取消，已完成 {len(results)} 条查询")
            break
        result = cache[name]
        results[name] = result
        skipped += 1
        completed_count += 1
        row_num = start_idx + company_names.index(name) + 1
        if emitter:
            emitter.progress("company_query", completed_count, total, f"[缓存] {name}", detail={"item": result, "row": row_num, "cached": True})

    if uncached_names and not is_cancelled(task_id):
        max_workers = max(1, config.get("max_concurrent", 3))
        _cancel_event = threading.Event()

        def _query_one(name: str) -> tuple[str, dict, int]:
            if _cancel_event.is_set():
                idx_in_list = company_names.index(name)
                row_num = start_idx + idx_in_list + 1
                return name, {"original_name": name, "current_name": "", "legal_person": "", "location": "", "credit_code": "", "status": "failed", "error": "任务已取消"}, row_num
            idx_in_list = company_names.index(name)
            row_num = start_idx + idx_in_list + 1
            if emitter:
                emitter.progress("company_query", completed_count + 1, total, f"查询: {name}")
            result = process_single_company(name, config)
            return name, result, row_num

        pending_futures = {}
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            name_iter = iter(uncached_names)
            for name in name_iter:
                if is_cancelled(task_id):
                    _cancel_event.set()
                    break
                future = executor.submit(_query_one, name)
                pending_futures[future] = name
                if len(pending_futures) >= max_workers:
                    break

            remaining_names = list(name_iter)

            while pending_futures:
                if is_cancelled(task_id):
                    _cancel_event.set()
                    for f in pending_futures:
                        f.cancel()
                    if emitter:
                        emitter.log("warn", f"任务已取消，已完成 {len(results)} 条查询")
                    for f in pending_futures:
                        try:
                            f.result(timeout=1)
                        except Exception:
                            pass
                    break

                done, _ = _wait_with_timeout(pending_futures, timeout=0.5)

                for future in done:
                    pending_futures.pop(future, None)
                    try:
                        name, result, row_num = future.result()
                    except Exception as e:
                        name = pending_futures.get(future, "")
                        result = {"original_name": name, "current_name": "", "legal_person": "", "location": "", "credit_code": "", "status": "failed", "error": str(e)}
                        row_num = 0

                    results[name] = result
                    completed_count += 1

                    if result["status"] == "success":
                        cache[name] = result
                        if completed_count % 5 == 0 or completed_count == total:
                            _save_cache(cache_path, list(cache.values()))

                    if result["status"] == "failed" and result.get("balance_depleted"):
                        balance_depleted = True
                        recharge_url = result.get("recharge_url", "")
                        _cancel_event.set()
                        if emitter:
                            emitter.log("warn", "API余额不足，停止后续查询。请充值后重试，已查询结果已缓存。")

                    if emitter:
                        emitter.progress("company_query", completed_count, total, f"完成: {name}", detail={"item": result, "row": row_num, "cached": False})

                    if remaining_names and not is_cancelled(task_id):
                        next_name = remaining_names.pop(0)
                        new_future = executor.submit(_query_one, next_name)
                        pending_futures[new_future] = next_name

    ordered_results = [results[name] for name in company_names if name in results]

    if balance_depleted:
        all_queried = set(results.keys())
        for name in company_names:
            if name not in all_queried:
                idx_in_list = company_names.index(name)
                row_num = start_idx + idx_in_list + 1
                results[name] = {
                    "original_name": name, "current_name": "", "legal_person": "",
                    "location": "", "credit_code": "", "status": "failed",
                    "error": "API余额不足，未查询",
                    "recharge_url": recharge_url,
                }
        ordered_results = [results[name] for name in company_names if name in results]

    _save_cache(cache_path, list(cache.values()))

    success_count = sum(1 for r in ordered_results if r["status"] == "success")
    warning_count = sum(1 for r in ordered_results if r["status"] == "warning")
    fail_count = sum(1 for r in ordered_results if r["status"] == "failed")
    cancelled = is_cancelled(task_id)

    if output_dir is None:
        from core.paths import USER_DATA_DIR
        output_dir = USER_DATA_DIR / "output" / "company-query"
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_excel_path = output_dir / f"企业查询结果_{timestamp}.xlsx"

    output_excel = _write_results_excel(ordered_results, output_excel_path, source_df=df, start_idx=start_idx)

    clear_cancel(task_id)

    return {
        "total": total,
        "success_count": success_count,
        "warning_count": warning_count,
        "fail_count": fail_count,
        "skipped_cached": skipped,
        "cancelled": cancelled,
        "balance_depleted": balance_depleted,
        "companies": ordered_results,
        "output_excel_path": output_excel,
    }
