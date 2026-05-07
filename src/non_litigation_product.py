#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


def load_non_litigation_cases(sample_root: Path) -> List[Dict]:
    excel_path = sample_root / '台账及命名规则.xlsx'
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active
    cases: List[Dict] = []
    for row in sheet.iter_rows(values_only=True):
        values = [str(item).strip() if item is not None else '' for item in row]
        if len(values) < 5:
            continue
        original_notice = values[1]
        renamed_notice = values[2]
        company_name = values[4]
        if '责字' not in original_notice or '责催-' not in renamed_notice or not company_name:
            continue
        sequence = renamed_notice.split('-责催-')[0].replace(' ', '')
        notice_number = original_notice.replace(' ', '')
        cases.append({
            'sequence': sequence,
            'notice_number': notice_number,
            'company_name': company_name,
        })
        if len(cases) == 3:
            break
    return cases


def build_non_litigation_standard_plan(sample_root: Path) -> Dict[str, List[Dict]]:
    cases = load_non_litigation_cases(sample_root)
    plan = {
        '责催': [],
        '申请书': [],
        '授权书': [],
        '所函': [],
    }
    for case in cases:
        sequence = case['sequence']
        notice_number = case['notice_number']
        company_name = case['company_name']
        plan['责催'].append({
            'target_filename': f'{sequence}-责催-{notice_number}.pdf',
            'company_name': company_name,
        })
        plan['申请书'].append({
            'target_filename': f'{sequence}-申请书pdf-{notice_number}.pdf',
            'company_name': company_name,
        })
        plan['授权书'].append({
            'target_filename': f'{company_name}.pdf',
            'company_name': company_name,
        })
        plan['所函'].append({
            'target_filename': f'{company_name}.pdf',
            'company_name': company_name,
        })
    return plan
