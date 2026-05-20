#!/usr/bin/env python3

import threading
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional

VIRTUAL_PRINTER_KEYWORDS = [
    "pdf", "xps", "onenote", "fax", "microsoft print to pdf",
    "导出为wps pdf", "adobe pdf", "foxit", "pdf24", "cutepdf",
    "bullzip", "primo", "pdfwriter", "image", "file",
]


def _is_virtual_printer(name: str) -> bool:
    name_lower = name.lower()
    return any(keyword in name_lower for keyword in VIRTUAL_PRINTER_KEYWORDS)


def list_printers(include_virtual: bool = False) -> List[dict]:
    try:
        import win32print
    except ImportError:
        return []

    physical_printers = []
    virtual_printers = []
    default_printer: Optional[str] = None

    try:
        default_printer = win32print.GetDefaultPrinter()
    except Exception:
        pass

    try:
        printer_flags = win32print.PRINTER_ENUM_LOCAL | win32print.PRINTER_ENUM_CONNECTIONS
        for printer_info in win32print.EnumPrinters(printer_flags):
            name = printer_info[2]
            is_virtual = _is_virtual_printer(name)
            try:
                handle = win32print.OpenPrinter(name)
                win32print.ClosePrinter(handle)
                entry = {
                    "name": name,
                    "is_default": name == default_printer,
                    "is_virtual": is_virtual,
                }
                if is_virtual:
                    virtual_printers.append(entry)
                else:
                    physical_printers.append(entry)
            except Exception:
                continue
    except Exception:
        pass

    if physical_printers:
        return physical_printers
    elif include_virtual or virtual_printers:
        return virtual_printers
    else:
        return []


def list_all_printers() -> List[dict]:
    return list_printers(include_virtual=True)


def get_default_printer() -> Optional[str]:
    try:
        import win32print
        return win32print.GetDefaultPrinter()
    except Exception:
        return None


def read_excel_columns(excel_path: str) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            if cell.value is not None:
                headers.append({"column": cell.column_letter, "name": str(cell.value).strip()})
        wb.close()
        return {"columns": headers, "count": len(headers)}
    except Exception as e:
        return {"columns": [], "count": 0, "error": str(e)}


def read_excel_column_values(excel_path: str, column_letter: str, range_start: int, range_end: int) -> dict:
    try:
        import openpyxl
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb.active
        values = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=range_start, max_row=range_end, min_col=_col_to_idx(column_letter), max_col=_col_to_idx(column_letter)), start=range_start):
            cell_val = row[0].value if row else None
            if cell_val is not None and str(cell_val).strip():
                values.append({"row": row_idx, "value": str(cell_val).strip()})
        wb.close()
        return {"values": values, "count": len(values)}
    except Exception as e:
        return {"values": [], "count": 0, "error": str(e)}


def _col_to_idx(letter: str) -> int:
    result = 0
    for ch in letter.upper():
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def _match_files_by_keyword(folder: Path, keyword: str) -> List[Path]:
    if not keyword:
        return []
    keyword_lower = keyword.lower()
    matched = []
    for p in folder.rglob("*"):
        if p.is_file() and p.suffix.lower() == ".pdf" and keyword_lower in p.name.lower():
            matched.append(p)
    matched.sort(key=lambda p: p.name)
    return matched


def _print_pdf_silent(pdf_path: Path, printer_name: str, copies: int = 1,
                      start_page: Optional[int] = None, end_page: Optional[int] = None) -> dict:
    try:
        import win32api
        import win32print

        try:
            handle = win32print.OpenPrinter(printer_name)
            win32print.ClosePrinter(handle)
        except Exception as e:
            return {"filename": pdf_path.name, "status": "failed", "error": f"打印机不可用: {str(e)}"}

        abs_path = str(pdf_path.resolve())

        result = win32api.ShellExecute(
            0,
            "print",
            abs_path,
            f'/d:"{printer_name}"',
            ".",
            0,
        )

        if result > 32:
            return {"filename": pdf_path.name, "status": "submitted", "message": "已提交"}
        else:
            error_map = {
                0: "内存不足", 2: "文件未找到", 3: "路径未找到",
                5: "访问拒绝", 31: "无关联程序", 32: "DLL未找到",
            }
            error_msg = error_map.get(result, f"未知错误码: {result}")
            return {"filename": pdf_path.name, "status": "failed", "error": error_msg}
    except Exception as e:
        return {"filename": pdf_path.name, "status": "failed", "error": str(e)}


@dataclass
class PrintTask:
    task_id: str
    status: str = "pending"
    total_jobs: int = 0
    completed_jobs: int = 0
    failed_jobs: int = 0
    current_file: str = ""
    current_company: str = ""
    printer_name: str = ""
    cancel_event: threading.Event = field(default_factory=threading.Event)
    started_at: Optional[float] = None
    finished_at: Optional[float] = None
    errors: List[dict] = field(default_factory=list)


class PrintTaskManager:
    _instance = None
    _lock = threading.Lock()

    def __new__(cls):
        with cls._lock:
            if cls._instance is None:
                cls._instance = super().__new__(cls)
                cls._instance._tasks: Dict[str, PrintTask] = {}
                cls._instance._current_task_id: Optional[str] = None
            return cls._instance

    def start_task(self, task_id: str) -> PrintTask:
        task = PrintTask(task_id=task_id)
        self._tasks[task_id] = task
        self._current_task_id = task_id
        return task

    def get_task(self, task_id: str) -> Optional[PrintTask]:
        return self._tasks.get(task_id)

    def get_current_task_id(self) -> Optional[str]:
        return self._current_task_id

    def cancel_task(self, task_id: str) -> bool:
        task = self._tasks.get(task_id)
        if task and task.status in ("pending", "running"):
            task.cancel_event.set()
            task.status = "cancelled"
            return True
        return False

    def finish_task(self, task_id: str, status: str = "completed"):
        task = self._tasks.get(task_id)
        if task:
            task.status = status
            task.finished_at = time.time()
            if self._current_task_id == task_id:
                self._current_task_id = None

    def cleanup_old_tasks(self, keep: int = 5):
        if len(self._tasks) <= keep:
            return
        finished = [(tid, t) for tid, t in self._tasks.items() if t.status in ("completed", "cancelled", "failed")]
        finished.sort(key=lambda x: x[1].finished_at or 0)
        for tid, _ in finished[:len(finished) - keep]:
            del self._tasks[tid]


_mgr = PrintTaskManager()


def process_print_v2(
    excel_path: str,
    folder_path: str,
    column_name: str,
    range_start: int,
    range_end: int,
    printer_name: str,
    copies: int = 1,
    page_start: Optional[int] = None,
    page_end: Optional[int] = None,
    print_mode: str = "single",
    dry_run: bool = False,
    selected_orders: Optional[List[int]] = None,
    task_id: str = "print-0",
    emitter=None,
) -> dict:
    task = _mgr.start_task(task_id)
    task.printer_name = printer_name
    task.status = "running"
    task.started_at = time.time()

    try:
        folder = Path(folder_path)
        if not folder.exists():
            raise FileNotFoundError(f"文件夹不存在: {folder}")

        if not dry_run and _is_virtual_printer(printer_name):
            dry_run = True
            if emitter:
                emitter.log("warn", f"检测到虚拟打印机「{printer_name}」，自动切换为匹配预览模式")

        if not printer_name:
            default = get_default_printer()
            if default:
                printer_name = default
                task.printer_name = printer_name

        company_entries = []
        if excel_path and column_name:
            col_data = read_excel_column_values(excel_path, column_name, range_start, range_end)
            company_entries = col_data.get("values", [])

        match_results: List[dict] = []
        order = 0

        if not company_entries:
            all_pdfs = sorted([p for p in folder.rglob("*") if p.is_file() and p.suffix.lower() == ".pdf"],
                              key=lambda p: p.name)
            if not all_pdfs:
                raise Exception(f"文件夹中无 PDF 文件: {folder}")

            for pdf_path in all_pdfs:
                order += 1
                match_results.append({
                    "order": order,
                    "company": "",
                    "row": 0,
                    "files": [{"name": pdf_path.name, "path": str(pdf_path.relative_to(folder))}],
                    "status": "matched",
                })
        else:
            for entry in company_entries:
                company_name = entry["value"]
                row_idx = entry["row"]
                matched = _match_files_by_keyword(folder, company_name)
                order += 1
                if matched:
                    match_results.append({
                        "order": order,
                        "company": company_name,
                        "row": row_idx,
                        "files": [{"name": p.name, "path": str(p.relative_to(folder))} for p in matched],
                        "status": "matched",
                    })
                else:
                    match_results.append({
                        "order": order,
                        "company": company_name,
                        "row": row_idx,
                        "files": [],
                        "status": "no_match",
                    })
                    task.errors.append({"company": company_name, "error": "未找到匹配文件"})
                    if emitter:
                        emitter.log("warn", f"行{row_idx}: 「{company_name}」未找到匹配PDF")

        total_matched = sum(len(m["files"]) for m in match_results if m["status"] == "matched")
        total_unmatched = sum(1 for m in match_results if m["status"] == "no_match")

        if selected_orders:
            order_set = set(selected_orders)
            match_results = [m for m in match_results if m["order"] in order_set]
            total_matched = sum(len(m["files"]) for m in match_results if m["status"] == "matched")
            total_unmatched = sum(1 for m in match_results if m["status"] == "no_match")

        if emitter:
            mode_label = "匹配预览" if dry_run else "打印"
            emitter.log("info", f"{mode_label}: {total_matched} 个PDF已匹配, {total_unmatched} 条记录无匹配, 打印机={printer_name}")

        if dry_run:
            _mgr.finish_task(task_id, "completed")
            return {
                "task_id": task_id,
                "status": "completed",
                "dry_run": True,
                "total_jobs": total_matched,
                "submitted": total_matched,
                "failed": total_unmatched,
                "printer_used": printer_name,
                "errors": task.errors[:50],
                "match_results": match_results,
            }

        task.total_jobs = total_matched
        submitted = 0
        failed = 0

        for match_entry in match_results:
            if match_entry["status"] != "matched":
                continue
            company_name = match_entry["company"]

            for file_info in match_entry["files"]:
                pdf_path = folder / file_info["path"]
                if not pdf_path.exists():
                    continue

                if task.cancel_event.is_set():
                    task.status = "cancelled"
                    break

                task.current_file = pdf_path.name
                task.current_company = company_name

                if emitter:
                    detail = {"company": company_name, "file": pdf_path.name}
                    emitter.progress("printing", submitted + failed + 1,
                                     task.total_jobs, f"打印: {pdf_path.name}", detail=detail)

                result = _print_pdf_silent(pdf_path, printer_name, copies, page_start, page_end)

                if result["status"] == "submitted":
                    submitted += 1
                    task.completed_jobs += 1
                else:
                    failed += 1
                    task.failed_jobs += 1
                    task.errors.append({"company": company_name, "file": pdf_path.name, "error": result.get("error", "")})

                time.sleep(0.3)

            if task.cancel_event.is_set():
                break

        _mgr.finish_task(task_id, "cancelled" if task.cancel_event.is_set() else "completed")

        if emitter:
            if failed > 0:
                emitter.log("warn", f"打印完成: {submitted} 成功, {failed} 失败")
            else:
                emitter.log("info", f"打印完成: {submitted} 个文件已提交")

        return {
            "task_id": task_id,
            "status": task.status,
            "dry_run": False,
            "total_jobs": task.total_jobs,
            "submitted": submitted,
            "failed": failed,
            "printer_used": printer_name,
            "errors": task.errors[:50],
            "match_results": match_results,
        }
    except Exception as e:
        _mgr.finish_task(task_id, "failed")
        if emitter:
            emitter.log("error", f"打印任务失败: {str(e)}")
        raise


def cancel_print_task(task_id: str) -> bool:
    return _mgr.cancel_task(task_id)


def get_print_task_status(task_id: str) -> Optional[dict]:
    task = _mgr.get_task(task_id)
    if not task:
        return None
    return {
        "task_id": task.task_id,
        "status": task.status,
        "total_jobs": task.total_jobs,
        "completed_jobs": task.completed_jobs,
        "failed_jobs": task.failed_jobs,
        "current_file": task.current_file,
        "current_company": task.current_company,
        "printer_name": task.printer_name,
        "started_at": task.started_at,
        "finished_at": task.finished_at,
        "error_count": len(task.errors),
    }


def check_printer_status(printer_name: str) -> dict:
    try:
        import win32print
        handle = win32print.OpenPrinter(printer_name)
        try:
            info = win32print.GetPrinter(handle, 2)
            status_code = info['Status']
            status_messages = []
            if status_code == 0:
                status_messages.append("就绪")
            if status_code & win32print.PRINTER_STATUS_PAUSED:
                status_messages.append("已暂停")
            if status_code & win32print.PRINTER_STATUS_ERROR:
                status_messages.append("错误")
            if status_code & win32print.PRINTER_STATUS_OFFLINE:
                status_messages.append("离线")
            if status_code & win32print.PRINTER_STATUS_OUT_OF_MEMORY:
                status_messages.append("内存不足")
            if status_code & win32print.PRINTER_STATUS_DOOR_OPEN:
                status_messages.append("打印机盖打开")
            if status_code & win32print.PRINTER_STATUS_NO_TONER:
                status_messages.append("无墨粉")
            if status_code & win32print.PRINTER_STATUS_PAPER_OUT:
                status_messages.append("缺纸")
            if status_code & win32print.PRINTER_STATUS_PAPER_JAM:
                status_messages.append("卡纸")
            return {
                "name": printer_name,
                "available": True,
                "status_code": status_code,
                "status": ", ".join(status_messages) if status_messages else "未知",
                "is_ready": status_code == 0,
            }
        finally:
            win32print.ClosePrinter(handle)
    except Exception as e:
        return {"name": printer_name, "available": False, "error": str(e)}
