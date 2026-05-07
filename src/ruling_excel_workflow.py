#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


EXPECTED_HEADERS = ['区号', '行政审查案号', '责令号', '被执行人', '职工姓名', '金额', '审批员/法官助理', '执行时间']

SOURCE_HEADER_MAPPING = {
    '区域': '区号',
    '责令号': '责令号',
    '被执行人': '被执行人',
    '职工': '职工姓名',
    '金额': '金额',
}


def load_non_litigation_rows_from_excel(excel_path: Path) -> List[Dict]:
    workbook = load_workbook(excel_path, data_only=True)
    sheet = workbook.active

    rows = list(sheet.iter_rows(values_only=True))
    header_row_index = None
    header_map: Dict[str, int] = {}

    for idx, row in enumerate(rows):
        values = [str(item).strip() if item is not None else '' for item in row]
        if '区域' in values and '责令号' in values and '被执行人' in values:
            header_row_index = idx
            header_map = {value: i for i, value in enumerate(values) if value}
            break

    if header_row_index is None:
        return []

    result: List[Dict] = []
    for row in rows[header_row_index + 1:]:
        values = [str(item).strip() if item is not None else '' for item in row]
        if not any(values):
            continue
        item: Dict[str, str] = {header: '' for header in EXPECTED_HEADERS}
        for source_header, target_header in SOURCE_HEADER_MAPPING.items():
            index = header_map.get(source_header)
            if index is not None and index < len(values):
                item[target_header] = values[index]
        if item.get('责令号'):
            result.append(item)

    return result
